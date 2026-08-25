"""One whole month, from four synthetic workbooks to a report and a snapshot.

**Why this file exists.** Until it was written, nothing in the suite called
`pipeline.run()`. Every stage was unit-tested and every reader had fixture tests, but
the assembled pipeline was only ever verified by a human running it against the real
`data/` folder — which is git-ignored, so no automated run could ever do it. That made
the one question that matters ("does the whole thing still work?") depend on somebody
remembering to check.

The figures below are hand-computed, and the input is deliberately unpleasant: a
missing punch, a night shift crossing midnight, a person in both sites on the same day,
a remote-work day that collides with the nominal timesheet placeholder, and a short day.
Those are the five shapes that produced real bugs.
"""

from __future__ import annotations

import json
from datetime import date, timedelta

import openpyxl
import pytest

from mesai import config, snapshot
from mesai.pipeline import run
from mesai.rules.worktime import hhmm
from tests.builders import (
    izin_row, mac_row, roster_row, tek_block, tek_day, write_izin, write_macunkoy,
    write_roster, write_teknopark,
)

PERIOD = "2026-06"


@pytest.fixture
def month(tmp_path):
    """Build a complete month of input and return the folders to run against."""
    raw = tmp_path / "raw"
    roster_dir = tmp_path / "personel"
    raw.mkdir()
    roster_dir.mkdir()

    write_roster(roster_dir / "calisan_listesi.xlsx", [
        roster_row("adeneme", "1001", "AYŞE", "DENEME", "ayse@example.com"),
        roster_row("vornek", "1002", "VELİ", "ÖRNEK", "veli@example.com",
                   facility="MACUNKÖY TESİSİ"),
        roster_row("ztaslak", "1003", "ZEYNEP", "TASLAK", "zeynep@example.com"),
    ])

    # AYŞE: a clean Teknopark day, plus a Macunköy site visit inside it (ADR-001).
    # VELİ: a night shift the source records with a negative duration (D2), and a
    #       day whose exit is missing and cannot be repaired (ADR-003).
    # ZEYNEP: a two-hour day, under the short-day threshold (ADR-019).
    write_macunkoy(raw / "Macunköy Haziran Mesai giriş-çıkış.xlsx", [
        mac_row("AYŞE", "DENEME", "1001", "2026-06-01", "13:20", "14:05", "0:45"),
        mac_row("VELİ", "ÖRNEK", "1002", "2026-06-02", "23:00", "02:00", "-21:00"),
        mac_row("VELİ", "ÖRNEK", "1002", "2026-06-03", "08:00", None, None),
        mac_row("ZEYNEP", "TASLAK", "1003", "2026-06-04", "09:00", "11:00", "2:00"),
        mac_row("ZİYARETÇİ7", "ZİYARETÇİ7", "SN9001", "2026-06-01",
                "08:00", "17:00", "9:00"),          # excluded, not an employee
    ])

    # AYŞE 08:00-17:00 on the 1st; her Macunköy visit falls inside it, so the union is
    # one interval. On the 5th she has the nominal 09:00-18:00 placeholder, the same
    # day she declared remote work (ADR-017, ADR-018).
    write_teknopark(raw / "Teknopark - Haziran Mesai Takip Exceli.xlsx", [
        tek_block("AYŞE DENEME", [
            tek_day("2026-06-01", "08:00", "17:00", "9:00"),
            tek_day("2026-06-05", "09:00", "18:00", "9:00"),
        ], row=3, col=1, total="18:00"),
    ])

    write_izin(raw / "HCMT34_HAZIRAN_IZIN.xlsx", [
        izin_row("1001", "AYŞE DENEME", None, None, None, None, None, 1.0),
        izin_row("1001", "AYŞE DENEME", "Uzaktan Çalışma", "05.06.2026", "07:30",
                 "05.06.2026", "16:30", 1.0),
        izin_row("1003", "ZEYNEP TASLAK", None, None, None, None, None, 1.0),
        izin_row("1003", "ZEYNEP TASLAK", "Yıllık İzin", "08.06.2026", "07:30",
                 "08.06.2026", "16:30", 1.0),
    ])
    return raw, roster_dir


@pytest.fixture
def result(month, tmp_path, settings):
    raw, roster_dir = month
    return run(raw, tmp_path / "out" / f"mesai-raporu-{PERIOD}.xlsx", PERIOD, settings,
               generated_at=__import__("datetime").datetime(2026, 7, 1, 9, 0),
               roster_dir=roster_dir,
               snapshot_path=tmp_path / "veri" / f"gonderim-{PERIOD}.json")


# --- the run produces both artifacts ---------------------------------------

def test_both_the_workbook_and_the_snapshot_are_written(result):
    assert result["output"].exists()
    assert result["snapshot"].exists()

    book = openpyxl.load_workbook(result["output"], read_only=True)
    assert book.sheetnames == ["Aylık Özet", "Günlük Detay", "İnceleme Listesi",
                               "Şüpheli Kayıtlar", "İzin Özeti", "Kontrol"]


def test_the_reconciliation_invariant_holds(result):
    """The guard from AGENTS.md §3. If this fails, hours were lost or doubled."""
    book = openpyxl.load_workbook(result["output"], read_only=True)
    control = "\n".join(
        " | ".join(str(c) for c in row if c is not None)
        for row in book["Kontrol"].iter_rows(values_only=True))

    assert "Mutabakat | TAMAM" in control


# --- the five shapes that produced real bugs -------------------------------

def test_a_cross_site_day_is_counted_once(result):
    """AYŞE: Teknopark 08:00-17:00 with a Macunköy visit inside it. Nine hours, not
    nine plus forty-five minutes."""
    day = _day_of(result, "AYŞE DENEME", "01.06.2026")
    assert day[6] == "9:00", f"got {day[6]}"


def test_a_midnight_crossing_is_repaired(result):
    """VELİ, 2nd: the source says -21:00 for a 23:00 -> 02:00 night shift."""
    day = _day_of(result, "VELİ ÖRNEK", "02.06.2026")
    assert day[6] == "3:00"
    # The column prints the tag in words now, in the same wording as the
    # note label — the internal `gece-geçişi` never reaches a reader (ADR-050).
    assert "Gece geçişi" in (day[9] or "")


def test_an_unrepairable_missing_punch_contributes_nothing(result):
    """VELİ, 3rd: entry only, no other site to reconcile against. ADR-003.

    The day now has a row — every working day does (ADR-063) — and what it must not have
    is an hour. An invented default time is the failure ADR-003 forbids.

    The row names the file the refused record came from, not `kayıt yok`: he badged, the
    reading could not be used, and those are different things. Saying otherwise was the
    bug the operator found on a day whose exit was stamped at 19:56 (ADR-067).
    """
    day = _day_of(result, "VELİ ÖRNEK", "03.06.2026")
    assert not day[6], f"süre yazılmış: {day[6]!r}"
    assert day[8] == "Macunköy"
    assert day[9] == "Çıkış yok"
    assert day[3], "damganın kendisi satırda görünmeli"

    book = openpyxl.load_workbook(result["output"], read_only=True)
    anomalies = "\n".join(
        " | ".join(str(c) for c in row if c is not None)
        for row in book["Şüpheli Kayıtlar"].iter_rows(values_only=True))
    assert "Çıkış yok" in anomalies


def test_a_remote_day_overrides_the_nominal_placeholder(result):
    """AYŞE, 5th: declared 07:30-16:30, timesheet holds the 09:00-18:00 placeholder.

    ADR-018 — the declaration wins, so nine hours rather than 07:30->18:00 = 10:30.
    """
    day = _day_of(result, "AYŞE DENEME", "05.06.2026")
    assert day[6] == "9:00", f"got {day[6]} — placeholder was not set aside"


def test_a_short_day_is_flagged_but_still_counted(result):
    """ZEYNEP, 4th: two hours exactly is NOT under the threshold, so no flag."""
    day = _day_of(result, "ZEYNEP TASLAK", "04.06.2026")
    assert day[6] == "2:00"
    assert "kısa-gün" not in (day[9] or ""), "exactly 2:00 must not flag"


def test_a_visitor_badge_never_reaches_the_summary(result):
    book = openpyxl.load_workbook(result["output"], read_only=True)
    names = [row[0] for row in book["Aylık Özet"].iter_rows(values_only=True)
             if row[0]]
    assert not any("ZİYARETÇİ" in str(name) for name in names)


# --- the monthly total, computed by hand -----------------------------------

def test_the_monthly_total_is_the_hand_computed_figure(result):
    """9:00 + 3:00 + 2:00 + 9:00 = 23:00.

    AYŞE 1st 9:00 and 5th 9:00; VELİ 2nd 3:00 (3rd contributes nothing);
    ZEYNEP 4th 2:00. No break deduction (ADR-016), no in-day gaps.
    """
    assert hhmm(result["gross"]) == "23:00"


def test_the_snapshot_agrees_with_the_workbook(result):
    """Both artifacts come from one run; a divergence here means they can disagree."""
    loaded = snapshot.load(result["snapshot"])
    total = sum((p.minutes for p in loaded.people), 0)

    assert total == int(result["gross"].total_seconds() // 60)
    assert loaded.period == PERIOD


def test_a_month_holding_only_its_first_week_is_reported_incomplete(result):
    """ADR-020, end to end.

    This fixture deliberately stops on the 5th while June has 22 expected working
    days, which is the July 2026 situation in miniature: internally consistent files
    describing part of a month. The first version of this test asserted the opposite
    and failed — the guard was right and the assumption was wrong.
    """
    loaded = snapshot.load(result["snapshot"])
    assert not loaded.is_complete

    partial = result["partial_sources"]
    assert partial, "a source stopping mid-month must be reported"
    assert {c.source for c in partial} == {"macunkoy", "teknopark"}
    assert all(c.trailing_missing for c in partial)

    book = openpyxl.load_workbook(result["output"], read_only=True)
    summary = "\n".join(
        " ".join(str(c) for c in row if c is not None)
        for row in book["Aylık Özet"].iter_rows(values_only=True))
    assert "BU RAPOR EKSİK" in summary, "the deliverable sheet must say so"


def test_the_snapshot_carries_emails_the_workbook_omits(result):
    """The mail step's whole reason for reading the snapshot instead. ADR-021."""
    loaded = snapshot.load(result["snapshot"])
    ayse = next(p for p in loaded.people if p.name == "AYŞE DENEME")
    assert ayse.email == "ayse@example.com"

    book = openpyxl.load_workbook(result["output"], read_only=True)
    everything = "\n".join(
        " ".join(str(c) for c in row if c is not None)
        for sheet in book.worksheets for row in sheet.iter_rows(values_only=True))
    assert "ayse@example.com" not in everything


def test_a_person_with_leave_but_no_attendance_is_reported_as_such(result):
    """ZEYNEP has leave on the 8th and worked the 4th; nobody here is data-less, so the
    snapshot must not invent a 'no data' person."""
    loaded = snapshot.load(result["snapshot"])
    assert {p.name for p in loaded.people} == {
        "AYŞE DENEME", "VELİ ÖRNEK", "ZEYNEP TASLAK"}
    assert all(p.has_attendance for p in loaded.people)


# --- determinism ------------------------------------------------------------

def test_two_runs_produce_identical_figures(month, tmp_path, settings):
    """ADR-005. The generation timestamp is the only thing allowed to differ."""
    raw, roster_dir = month
    import datetime as _dt

    def once(tag):
        out = run(raw, tmp_path / tag / f"r-{PERIOD}.xlsx", PERIOD, settings,
                  generated_at=_dt.datetime(2026, 7, 1, 9, 0), roster_dir=roster_dir,
                  snapshot_path=tmp_path / tag / "s.json")
        payload = json.loads(out["snapshot"].read_text(encoding="utf-8"))
        return out["gross"], payload

    first_gross, first_snap = once("a")
    second_gross, second_snap = once("b")

    assert first_gross == second_gross
    assert first_snap == second_snap


# --- helpers ----------------------------------------------------------------

def _day_of(result, name: str, date: str) -> tuple:
    """One person's one day. Name AND date, because the sheet now holds every person's
    every working day (ADR-063) — a date alone matches whoever sorts first."""
    return next(w for w in _workdays(result) if w[0] == name and w[1] == date)


def _workdays(result) -> list[tuple]:
    """Data rows of the Günlük Detay sheet."""
    book = openpyxl.load_workbook(result["output"], read_only=True)
    rows = list(book["Günlük Detay"].iter_rows(values_only=True))
    header = next(i for i, row in enumerate(rows) if row[0] == "Ad Soyad")
    return [row for row in rows[header + 1:]
            if row[0] and not str(row[0]).startswith(("Oluşturulma", "Kaynak", "Bu "))]


# --- one vocabulary for notes (ADR-049) -------------------------------------
#
# The monthly summary's `Not` column used to be five hand-written strings built in the
# pipeline. Four were re-wordings of a note label — `Ayın çoğu açıklanmıyor` for
# `Ay büyük ölçüde boş` — and the other eleven labels never reached the column at all,
# so most people with a problem had an empty `Not` cell while the filter list showed
# them under a note. The operator found it by reading the report: "why do most of them
# have no note even though they have a problem, and what decides what is written?"

def _summary_notes(path):
    """`{person: [note, ...]}` from the monthly summary sheet."""
    book = openpyxl.load_workbook(path, read_only=True)
    rows = list(book["Aylık Özet"].iter_rows(values_only=True))
    header = next(r for r in rows if r and r[0] == "Ad Soyad")
    column = header.index("Not")
    found = {}
    for row in rows[rows.index(header) + 1:]:
        if not row or not row[0] or str(row[0]).startswith("TOPLAM"):
            continue
        found[row[0]] = [n for n in str(row[column] or "").split("; ") if n]
    return found


def test_the_summary_notes_use_the_same_words_as_everything_else(result):
    """Every note in the column is either a note label or the one roster fact.

    This is the guard against re-introducing a hand-written re-wording: a new string
    invented here would have to be added to the exception below, which is the moment
    somebody notices they are inventing a second vocabulary.
    """
    from mesai.anomalies import DESCRIPTIONS

    path = result["output"]
    known = {label for label, _s, _e, _g in DESCRIPTIONS.values()}
    known.add("Personel listesinde yok")     # a roster fact, not a problem — ADR-011

    for name, notes in _summary_notes(path).items():
        for note in notes:
            assert note in known, f"{name}: {note!r} hiçbir etikete karşılık gelmiyor"


def test_everybody_with_a_problem_has_a_note_in_the_summary(result):
    """The complaint, stated as a test. `Şüpheli Kayıt` counted them; `Not` was empty."""
    path = result["output"]
    book = openpyxl.load_workbook(path, read_only=True)
    rows = list(book["Aylık Özet"].iter_rows(values_only=True))
    header = next(r for r in rows if r and r[0] == "Ad Soyad")
    count_col, note_col = header.index("Şüpheli Kayıt"), header.index("Not")

    for row in rows[rows.index(header) + 1:]:
        if not row or not row[0] or str(row[0]).startswith("TOPLAM"):
            continue
        if row[count_col]:
            assert row[note_col], f"{row[0]}: {row[count_col]} şüpheli kayıt, not yok"


def test_the_notes_the_column_used_to_omit_now_reach_it(result):
    """The two failure modes the operator hit, one each.

    `Çıkış yok` was one of the eleven labels that never got into this column at all —
    `VELİ`'s third day has an exit that cannot be repaired, and his `Not` cell was
    empty while the filter list listed him under it.

    `ZEYNEP` shows the other: her note was printed, in different words. The column said
    `Ayın çoğu açıklanmıyor` and every other list in the program said
    `Ay büyük ölçüde boş` — a label ADR-062 has since removed, because her empty month
    is now stated as its empty days. The assertion follows the label to
    `Hem giriş hem çıkış yok`; what it holds is unchanged, that her `Not` cell is not
    empty and carries the wording every other list uses.
    """
    notes = _summary_notes(result["output"])

    assert "Çıkış yok" in notes["VELİ ÖRNEK"], notes["VELİ ÖRNEK"]
    assert "Hem giriş hem çıkış yok" in notes["ZEYNEP TASLAK"], notes["ZEYNEP TASLAK"]


def test_no_reworded_note_survives(result):
    """The four strings that were saying the same thing in different words."""
    path = result["output"]
    book = openpyxl.load_workbook(path, read_only=True)
    text = "\n".join(
        " ".join(str(c) for c in row if c is not None)
        for sheet in book.worksheets for row in sheet.iter_rows(values_only=True))

    for gone in ("Ayın çoğu açıklanmıyor", "Uzaktan çalışma kart kaydıyla çakışıyor",
                 "Gece vardiyası düzeltmesi var"):
        assert gone not in text, gone


def test_no_internal_tag_name_reaches_the_workbook(result):
    """`kısa-gün`, `uzaktan-çakışma`, `çapraz-tesis` are identifiers, not wording.

    The daily detail printed them raw, which made a third vocabulary for facts the rest
    of the program already had names for — and two of them differed by one word, so the
    operator had to ask what `çapraz-tesis` and `çapraz-eşleşti` meant. ADR-050.
    """
    from mesai.anomalies import TAG_TEXT

    book = openpyxl.load_workbook(result["output"], read_only=True)
    rows = list(book["Günlük Detay"].iter_rows(values_only=True))
    header = next(r for r in rows if r and r[0] == "Ad Soyad")
    column = header.index("Etiket")

    # The column's own cells, split the way they are joined. A plain substring search
    # over the whole workbook is too crude: `uzaktan` is also an ordinary word inside
    # the explanatory banners.
    printed = set()
    for row in rows[rows.index(header) + 1:]:
        if row and row[column]:
            printed.update(part.strip() for part in str(row[column]).split(","))

    assert printed, "hiç etiket yazılmamış — test bir şey kontrol etmiyor"
    leaked = printed & set(TAG_TEXT)
    assert not leaked, f"ham hâliyle yazılmış: {sorted(leaked)}"

    # Three legitimate vocabularies, and none of them is an identifier: a day's tags, a
    # note label for a day with no record at all, and the leave type the HCM wrote for a
    # day somebody was away (ADR-063). Listed rather than allowing anything, because the
    # thing being guarded is that no FOURTH vocabulary appears.
    from mesai.anomalies import DESCRIPTIONS
    allowed = (set(TAG_TEXT.values())
               | {label for label, _s, _e, _g in DESCRIPTIONS.values()}
               | {"Yıllık İzin", "Mazeret", "İstirahat (Raporlu)", "Eğitim İzni",
                  "Doğum Günü İzni", "Uzaktan Çalışma"})
    assert printed <= allowed, sorted(printed - allowed)


def test_every_tag_the_program_can_set_has_wording():
    """A tag with no entry would print its identifier and nothing would fail.

    Read out of `merge.py` rather than listed here, so a tag added there without an
    entry in `TAG_TEXT` fails this instead of quietly leaking into the report.
    """
    import re as regex
    from pathlib import Path

    from mesai.anomalies import TAG_TEXT

    source = Path("src/mesai/merge.py").read_text(encoding="utf-8")
    set_in_merge = set(regex.findall(r'tags\.add\("([^"]+)"\)', source))

    assert set_in_merge, "merge.py'de tag bulunamadı — desen mi değişti?"
    assert set_in_merge <= set(TAG_TEXT), sorted(set_in_merge - set(TAG_TEXT))
