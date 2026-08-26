"""The report writer actually writes.

This file exists because of a real bug: the report kept its own copy of the
severity -> impact-text mapping, so adding the `info` severity in `anomalies.py`
(ADR-017) made every month-end run die with `KeyError: 'info'` — while all 112 unit
tests stayed green, because none of them built a workbook.

These tests are deliberately shallow. They do not check business figures; the merge
and worktime suites do that. They check that a workbook can be produced from every
severity and every anomaly kind, so a change to that vocabulary cannot pass CI and
fail on real data.
"""

from datetime import date, datetime, timedelta

import openpyxl
import pytest

from mesai.anomalies import IMPACT_TEXT, Anomaly, AnomalyKind, Collector, DESCRIPTIONS
from mesai.models import (
    Employee, Interval, LeaveRecord, MonthSummary, RunStats, WorkDay,
)
from mesai.report import workbook

KEY = ("AYSE", "DENEME")
DAY = date(2026, 5, 21)


def _employee() -> Employee:
    return Employee(
        key=KEY, display_name="AYŞE DENEME", personnel_no="8801",
        department="TEST EKİBİ", job_title="TEST MÜHENDİSİ",
        facility="DEICO TESİS", in_roster=True, sources=frozenset({"teknopark"}),
    )


def _workday() -> WorkDay:
    intervals = (
        Interval(datetime(2026, 5, 21, 8, 21), datetime(2026, 5, 21, 13, 48),
                 frozenset({"teknopark"})),
        Interval(datetime(2026, 5, 21, 14, 30), datetime(2026, 5, 21, 18, 0),
                 frozenset({"teknopark"})),
    )
    return WorkDay(
        key=KEY, date=DAY, intervals=intervals,
        gross=timedelta(hours=9, minutes=39), break_deduction=timedelta(),
        net=timedelta(hours=9, minutes=39), tags=frozenset({"uzaktan"}),
    )


def _summary() -> MonthSummary:
    return MonthSummary(
        employee=_employee(), period="2026-05",
        gross=timedelta(hours=9, minutes=39), net=timedelta(hours=9, minutes=39),
        worked_days=1, remote_days=1.0, leave_days=0.5, anomaly_count=0,
        has_attendance=True,
    )


def _all_kinds() -> Collector:
    """One anomaly of every kind — so no kind can be unrenderable."""
    collector = Collector()
    for row, kind in enumerate(AnomalyKind, start=1):
        collector.add(Anomaly(
            kind=kind, source="teknopark", source_row=row, key=KEY,
            raw_name="AYŞE DENEME", date=DAY, raw_entry="08:00", raw_exit="17:00",
            detail="test",
        ))
    return collector


def _build(tmp_path, settings, anomalies=None, summaries=None, name="rapor.xlsx"):
    path = tmp_path / name
    workbook.build(
        path=path, period="2026-05",
        summaries=summaries if summaries is not None else [_summary()],
        workdays=[_workday()], employees={KEY: _employee()},
        leave=[LeaveRecord(
            key=KEY, raw_name="AYŞE DENEME", personnel_no="8801",
            leave_type="Yıllık İzin", status="Kullanıldı",
            start=datetime(2026, 5, 20, 7, 30), end=datetime(2026, 5, 20, 12, 0),
            days=0.5, department="TEST EKİBİ", source_row=3)],
        anomalies=anomalies if anomalies is not None else _all_kinds(),
        stats=RunStats(
            rows_read={"teknopark": 1}, records_built={"teknopark": 1},
            intervals_accepted=2, union_total=timedelta(hours=8, minutes=57),
            accepted_total=timedelta(hours=9, minutes=39),
            files={"teknopark": "test.xlsx"}, roster_date=date(2026, 7, 28)),
        settings=settings, generated_at=datetime(2026, 8, 17, 12, 0),
    )
    return path


def test_every_anomaly_kind_can_be_written(tmp_path, settings):
    """The KeyError this file was written for. Covers all kinds and all severities."""
    path = _build(tmp_path, settings)
    assert path.exists()

    book = openpyxl.load_workbook(path, read_only=True)
    assert book.sheetnames == ["Aylık Özet", "Günlük Detay", "İnceleme Listesi",
                               "Şüpheli Kayıtlar", "İzin Özeti", "Kontrol"]


def test_every_severity_has_impact_text():
    """A kind whose severity has no impact text kills the run at report time."""
    for kind, (_, severity, _explanation, _group) in DESCRIPTIONS.items():
        assert severity in IMPACT_TEXT, f"{kind} has severity {severity!r}"


def test_every_anomaly_kind_is_described():
    for kind in AnomalyKind:
        assert kind in DESCRIPTIONS, f"{kind} would raise KeyError on .label"


def test_hours_columns_follow_the_break_switch(tmp_path, settings, settings_break):
    """ADR-016: one hours column pair with the deduction off, two with it on."""
    rows = openpyxl.load_workbook(_build(tmp_path, settings), read_only=True)
    headers = [c for c in list(rows["Aylık Özet"].iter_rows(values_only=True))[3]
               if c is not None]
    assert "Çalışma Süresi" in headers
    assert "Net Süre" not in headers

    book = openpyxl.load_workbook(
        _build(tmp_path, settings_break, name="eski-kural.xlsx"), read_only=True)
    headers = [c for c in list(book["Aylık Özet"].iter_rows(values_only=True))[3]
               if c is not None]
    assert "Brüt Süre" in headers and "Net Süre" in headers


def test_daily_sheet_shows_the_gap_that_the_envelope_pays(tmp_path, settings):
    book = openpyxl.load_workbook(_build(tmp_path, settings), read_only=True)
    rows = list(book["Günlük Detay"].iter_rows(values_only=True))
    headers = [c for c in rows[3] if c is not None]
    assert headers[6:8] == ["Çalışma Süresi", "Gün İçi Boşluk"]

    # By date, not by row number: the sheet holds every working day of the period now
    # (ADR-063), so the measured day is no longer the first row.
    day = next(r for r in rows[4:] if r and r[1] == DAY.strftime("%d.%m.%Y"))
    assert day[6] == "9:39" and day[7] == "0:42"


def test_the_rule_banner_names_the_active_rule(tmp_path, settings, settings_break):
    book = openpyxl.load_workbook(_build(tmp_path, settings), read_only=True)
    banner = list(book["Aylık Özet"].iter_rows(values_only=True))[2][0]
    assert "kesinti UYGULANMAZ" in banner
    assert "son çıkışına" in banner

    book = openpyxl.load_workbook(
        _build(tmp_path, settings_break, name="eski-kural.xlsx"), read_only=True)
    banner = list(book["Aylık Özet"].iter_rows(values_only=True))[2][0]
    assert "45 dk kesinti uygulanır" in banner


def test_locked_output_file_reports_a_readable_error(tmp_path, settings, monkeypatch):
    """It happens every month: HR leaves the workbook open in Excel."""
    def boom(*_args):
        raise PermissionError

    monkeypatch.setattr(workbook.os, "replace", boom)
    with pytest.raises(workbook.ReportLocked, match="kilitli"):
        _build(tmp_path, settings)


# --- partial-export banner (ADR-020) ---------------------------------------

def test_partial_coverage_puts_a_warning_above_the_table(tmp_path, settings):
    """The deliverable sheet must say so — a mid-month report looks normal otherwise."""
    from mesai.models import SourceCoverage
    stats = RunStats(
        rows_read={"teknopark": 1}, records_built={"teknopark": 1},
        intervals_accepted=2, union_total=timedelta(hours=8, minutes=57),
        accepted_total=timedelta(hours=9, minutes=39),
        files={"teknopark": "test.xlsx"}, roster_date=date(2026, 7, 28),
        coverage={"teknopark": SourceCoverage(
            source="teknopark", present=13, expected=23,
            trailing_missing=tuple(date(2026, 7, d) for d in (20, 21, 22)))},
    )
    path = tmp_path / "kismi.xlsx"
    workbook.build(
        path=path, period="2026-07", summaries=[_summary()], workdays=[_workday()],
        employees={KEY: _employee()}, leave=[], anomalies=Collector(), stats=stats,
        settings=settings, generated_at=datetime(2026, 8, 18, 12, 0),
    )

    book = openpyxl.load_workbook(path, read_only=True)
    rows = list(book["Aylık Özet"].iter_rows(values_only=True))
    assert "BU RAPOR EKSİK" in str(rows[3][0])
    assert "20.07" in str(rows[3][0])
    # The table must have shifted down rather than been overwritten.
    assert rows[4][0] == "Ad Soyad"

    control = "\n".join(
        " ".join(str(c) for c in r if c is not None)
        for r in book["Kontrol"].iter_rows(values_only=True))
    assert "KISMİ DIŞA AKTARIM" in control
    assert "13 / 23" in control


def test_full_coverage_leaves_the_table_where_it_was(tmp_path, settings):
    from mesai.models import SourceCoverage
    stats = RunStats(
        files={"teknopark": "test.xlsx"},
        coverage={"teknopark": SourceCoverage("teknopark", 23, 23, ())},
    )
    path = tmp_path / "tam.xlsx"
    workbook.build(
        path=path, period="2026-07", summaries=[_summary()], workdays=[_workday()],
        employees={KEY: _employee()}, leave=[], anomalies=Collector(), stats=stats,
        settings=settings, generated_at=datetime(2026, 8, 18, 12, 0),
    )

    book = openpyxl.load_workbook(path, read_only=True)
    rows = list(book["Aylık Özet"].iter_rows(values_only=True))
    assert rows[3][0] == "Ad Soyad", "no banner, header stays on row 4"
    control = "\n".join(
        " ".join(str(c) for c in r if c is not None)
        for r in book["Kontrol"].iter_rows(values_only=True))
    assert "Dönemin tamamı kapsanıyor" in control




def test_the_control_sheet_lists_only_this_month_s_holidays(tmp_path, settings):
    """July's report listed May's seven holidays under "assumptions behind the figures".

    The calendar file holds the whole year, so every month's report showed every
    month's holidays. Dates outside the period being reported on are not assumptions
    behind its figures; they are noise on the one sheet whose job is to be checkable.
    """
    from dataclasses import replace
    from datetime import date

    calendar = replace(
        settings.calendar,
        holidays=frozenset({date(2026, 5, 1), date(2026, 7, 15)}))
    path = tmp_path / "tatil.xlsx"
    workbook.build(
        path=path, period="2026-07", summaries=[_summary()], workdays=[_workday()],
        employees={KEY: _employee()}, leave=[], anomalies=Collector(),
        stats=RunStats(files={"teknopark": "test.xlsx"}),
        settings=replace(settings, calendar=calendar),
        generated_at=datetime(2026, 8, 20, 12, 0),
    )

    control = "\n".join(
        " ".join(str(c) for c in r if c is not None)
        for r in openpyxl.load_workbook(path, read_only=True)["Kontrol"].iter_rows(
            values_only=True))
    assert "15.07.2026" in control, "the month's own holiday must be listed"
    assert "01.05.2026" not in control, "another month's must not"


def test_a_month_with_no_holiday_defined_says_so(tmp_path, settings):
    """Silence would read as "there were none", which is the mistake to avoid.

    Ramazan and Kurban Bayramı move every year and are entered by hand, so a month
    that has nothing in the calendar is exactly the case worth naming out loud.
    """
    from dataclasses import replace

    path = tmp_path / "tatilsiz.xlsx"
    workbook.build(
        path=path, period="2026-07", summaries=[_summary()], workdays=[_workday()],
        employees={KEY: _employee()}, leave=[], anomalies=Collector(),
        stats=RunStats(files={"teknopark": "test.xlsx"}),
        settings=replace(settings,
                         calendar=replace(settings.calendar,
                                          holidays=frozenset())),
        generated_at=datetime(2026, 8, 20, 12, 0),
    )

    control = "\n".join(
        " ".join(str(c) for c in r if c is not None)
        for r in openpyxl.load_workbook(path, read_only=True)["Kontrol"].iter_rows(
            values_only=True))
    assert "işaretli gün yok" in control


def test_the_control_sheet_sections_are_numbered_once_each(tmp_path, settings):
    """Two sections were both numbered 9, and a doc pointed at the wrong one.

    The numbers are how everything else refers to this sheet — "section 7 lists every
    alias in effect" is the documented way to check that the alias table loaded at all.
    A duplicate makes such a reference ambiguous, and it was invisible because nothing
    read the numbers back.
    """
    import re

    path = tmp_path / "numaralar.xlsx"
    workbook.build(
        path=path, period="2026-07", summaries=[_summary()], workdays=[_workday()],
        employees={KEY: _employee()}, leave=[], anomalies=Collector(),
        stats=RunStats(files={"teknopark": "test.xlsx"}), settings=settings,
        generated_at=datetime(2026, 8, 20, 12, 0),
    )

    book = openpyxl.load_workbook(path, read_only=True)
    numbers = [int(m.group(1)) for r in book["Kontrol"].iter_rows(values_only=True)
               if r and r[0] and (m := re.match(r"^(\d+)\. ", str(r[0])))]

    assert numbers, "the sheet must have numbered sections at all"
    assert numbers == sorted(set(numbers)), f"repeated or out of order: {numbers}"
    assert numbers == list(range(1, len(numbers) + 1)), f"a gap: {numbers}"


# --- the workbook is for HR, not for developers -----------------------------

_DEVELOPER_JARGON = (
    "ADR-", "ROADMAP", "DOMAIN-RULES", "DATA-SOURCES", "OUTPUT-SPEC",
    ".md", ".py", "config/", "daily_hours", "break.deduct", "nominal_day",
    "Faz 2", "Faz 4",
)


def test_no_developer_references_reach_the_workbook(tmp_path, settings):
    """The person who reads this file will never open the repository.

    Cells used to carry things like `ROADMAP.md Q4` and `— ADR-015`: meaningful to
    whoever wrote the rule, noise to an HR reader who cannot follow them anywhere.
    48 cells in the May 2026 report were like this. Every explanation now stands on
    its own in plain Turkish.

    This test fails on the next `ADR-0NN` someone appends to a report string.
    """
    from mesai.models import SourceCoverage
    stats = RunStats(
        rows_read={"teknopark": 1}, records_built={"teknopark": 1},
        intervals_accepted=2, union_total=timedelta(hours=8, minutes=57),
        accepted_total=timedelta(hours=9, minutes=39),
        files={"teknopark": "test.xlsx"}, roster_date=date(2026, 7, 28),
        roster_duplicates=["AYŞE DENEME: iki hesap"],
        out_of_period={"teknopark": 2}, out_of_period_leave=1,
        # Both branches, so partial-coverage wording is covered too.
        coverage={"teknopark": SourceCoverage("teknopark", 13, 23,
                                             (date(2026, 7, 20), date(2026, 7, 21))),
                  "macunkoy": SourceCoverage("macunkoy", 23, 23, ())},
    )
    path = tmp_path / "jargon.xlsx"
    workbook.build(
        path=path, period="2026-07", summaries=[_summary()], workdays=[_workday()],
        employees={KEY: _employee()}, leave=[], anomalies=_all_kinds(), stats=stats,
        settings=settings, generated_at=datetime(2026, 8, 18, 12, 0),
    )

    offenders: list[str] = []
    book = openpyxl.load_workbook(path, read_only=True)
    for sheet in book.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if not isinstance(cell, str):
                    continue
                for token in _DEVELOPER_JARGON:
                    if token in cell:
                        offenders.append(f"{sheet.title}: {token!r} in {cell[:70]!r}")

    assert not offenders, "developer jargon in the workbook:\n" + "\n".join(offenders)


# --- where the day started and where it ended -------------------------------

def _day_label(workday, settings):
    from mesai.report.workbook import _day_sources_label
    return _day_sources_label(workday)


def test_a_day_at_one_site_is_named_plainly(settings):
    from mesai.models import Interval, WorkDay

    day = WorkDay(key=KEY, date=date(2026, 7, 1),
                  intervals=(Interval(datetime(2026, 7, 1, 8),
                                      datetime(2026, 7, 1, 17),
                                      frozenset({"teknopark"})),),
                  gross=timedelta(hours=9), break_deduction=timedelta(),
                  net=timedelta(hours=9))

    assert _day_label(day, settings) == "Teknopark"


def test_a_day_that_starts_at_one_site_and_ends_at_the_other_says_so(settings):
    """The operator's request: first punch at Macunköy, last at Teknopark — say it.

    Measured on July 2026: 8 person-days out of 2 731. Rare, and impossible to work out
    from the old column, which printed the union — `Macunköy + Teknopark` — for both
    this and the merged case below.
    """
    from mesai.models import Interval, WorkDay

    day = WorkDay(key=KEY, date=date(2026, 7, 1),
                  intervals=(Interval(datetime(2026, 7, 1, 8),
                                      datetime(2026, 7, 1, 9),
                                      frozenset({"macunkoy"})),
                             Interval(datetime(2026, 7, 1, 10),
                                      datetime(2026, 7, 1, 17),
                                      frozenset({"teknopark"}))),
                  gross=timedelta(hours=9), break_deduction=timedelta(),
                  net=timedelta(hours=9))

    assert _day_label(day, settings) == "Macunköy → Teknopark"


def test_a_merged_interval_is_still_shown_with_a_plus(settings):
    """284 of July's cross-site days are this shape, and the arrow would be a lie.

    When the first interval itself carries both sites, the day did not start at one of
    them — the two records overlap. Saying `Macunköy → Teknopark` would invent an order
    that the data does not have.
    """
    from mesai.models import Interval, WorkDay

    day = WorkDay(key=KEY, date=date(2026, 7, 1),
                  intervals=(Interval(datetime(2026, 7, 1, 8),
                                      datetime(2026, 7, 1, 17),
                                      frozenset({"macunkoy", "teknopark"})),),
                  gross=timedelta(hours=9), break_deduction=timedelta(),
                  net=timedelta(hours=9))

    assert _day_label(day, settings) == "Macunköy + Teknopark"


def test_a_day_with_no_interval_says_nothing(settings):
    from mesai.models import WorkDay

    day = WorkDay(key=KEY, date=date(2026, 7, 1), intervals=(),
                  gross=timedelta(), break_deduction=timedelta(), net=timedelta())

    assert _day_label(day, settings) == ""


# --- the workbook does not name who to ask ----------------------------------

# Whole words only. `İK` is a substring of ordinary Turkish — `EKSİK`, `İKİ` — so a
# naive `in` check would fail on the report's own warnings.
_DEPARTMENTS = ("İK", "IK", "IT", "HR")

# Phrases that hand the reader off to somebody, or credit a decision to somebody.
# Matched as substrings, which is safe because none of them is a fragment of an
# ordinary word. Lower-cased before comparison.
_HANDOFFS = (
    "onay bekl",          # "İK onayı bekliyor"
    "onayı bekl",
    "talebiyle",          # "45 dk kesinti İK talebiyle kapatıldı"
    "talimatıyla",
    "isteğiyle",
    "sorulacak",          # sheet 3 was called "Sorulacaklar"
    "sorulmalı",
    "sormak için",
    "ile kontrol edilmeli",   # "İK/IT ile kontrol edilmeli"
)


def _tokens(text: str) -> set[str]:
    """Upper-case words, with a Turkish suffix apostrophe cut off (`İK'ya` -> `İK`)."""
    import re

    words = re.split(r"[^0-9A-Za-zÇĞİÖŞÜçğıöşü']+", text)
    return {word.split("'", 1)[0].upper() for word in words if word}


def test_the_workbook_never_says_who_to_ask(tmp_path, settings):
    """It said `İK talebiyle kapatıldı`, `İK onayı bekliyor`, `İK'ya / IT'ye sormak`.

    The operator who runs this monthly is often not in contact with HR at all, so those
    lines described a process that was not happening and put words in somebody's mouth.
    A report states what was done and what is unresolved; who to take it to is not its
    business, and naming the wrong department is worse than naming none.

    Same shape as the developer-jargon test above and for the same reason: the reader
    cannot follow such a pointer anywhere.

    **This is about the program's own wording, not about the data.** Real departments
    and job titles contain the word — `İK VE OPERASYONLAR EKİBİ`, `... İK MÜDÜRÜ` — and
    those are the roster's text passed straight through, so cells carrying them are
    skipped rather than the check being weakened.
    """
    from mesai.models import SourceCoverage

    employee = _employee()
    passed_through = {employee.department, employee.job_title, employee.facility,
                      employee.display_name}

    stats = RunStats(
        rows_read={"teknopark": 1}, records_built={"teknopark": 1},
        intervals_accepted=2, union_total=timedelta(hours=8, minutes=57),
        accepted_total=timedelta(hours=9, minutes=39),
        files={"teknopark": "test.xlsx"}, roster_date=date(2026, 7, 28),
        roster_duplicates=["AYŞE DENEME: iki hesap"],
        out_of_period={"teknopark": 2}, out_of_period_leave=1,
        coverage={"teknopark": SourceCoverage("teknopark", 13, 23,
                                             (date(2026, 7, 20), date(2026, 7, 21))),
                  "macunkoy": SourceCoverage("macunkoy", 23, 23, ())},
    )
    path = tmp_path / "kim.xlsx"
    workbook.build(
        path=path, period="2026-07", summaries=[_summary()], workdays=[_workday()],
        employees={KEY: employee}, leave=[], anomalies=_all_kinds(), stats=stats,
        settings=settings, generated_at=datetime(2026, 8, 21, 12, 0),
    )

    offenders: list[str] = []
    book = openpyxl.load_workbook(path, read_only=True)
    names = [sheet.title for sheet in book.worksheets]
    for sheet in book.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if not isinstance(cell, str) or cell in passed_through:
                    continue
                for word in _tokens(cell) & set(_DEPARTMENTS):
                    offenders.append(f"{sheet.title}: {word!r} in {cell[:70]!r}")
                for phrase in _HANDOFFS:
                    if phrase in cell.lower():
                        offenders.append(
                            f"{sheet.title}: {phrase!r} in {cell[:70]!r}")

    # Sheet names are wording too — sheet 3 was called `Sorulacaklar`.
    for name in names:
        for phrase in _HANDOFFS:
            if phrase in name.lower():
                offenders.append(f"sayfa adı: {phrase!r} in {name!r}")
        for word in _tokens(name) & set(_DEPARTMENTS):
            offenders.append(f"sayfa adı: {word!r} in {name!r}")

    assert not offenders, ("the workbook names somebody or hands off to them:\n"
                           + "\n".join(offenders))


def test_the_check_would_catch_the_wording_that_was_removed():
    """A guard nobody has seen fail is a guard nobody should trust.

    Also pins the whole-word rule: the report's own `BU RAPOR EKSİK` contains `İK` and
    must not trip it.
    """
    assert _tokens("45 dk kesinti İK talebiyle kapatıldı") & set(_DEPARTMENTS)
    assert _tokens("Bu sayfa İK'ya / IT'ye sormak için hazırlanmıştır") & \
        set(_DEPARTMENTS)
    assert _tokens("Bu kişilerin ayı eksik — İK/IT ile kontrol edilmeli") & \
        set(_DEPARTMENTS)

    assert not _tokens("BU RAPOR EKSİK — kaynak dosyalar dönemin tamamını içermiyor") \
        & set(_DEPARTMENTS)
    # A real department, which the roster writes and the report only passes on.
    assert _tokens("İK VE OPERASYONLAR EKİBİ") & set(_DEPARTMENTS), \
        "the token check does see it — which is why such cells are skipped by name"
    assert not _tokens("İKİ kişinin saatleri birleşmiş olur") & set(_DEPARTMENTS)
    assert not _tokens("Mola için süre düşülmedi") & set(_DEPARTMENTS)

    # The handoff phrases, which is the other half of the rule.
    assert any(p in "aşağıdaki eşleştirmeler uygulandı ama i̇k onayı bekliyor"
               for p in _HANDOFFS)
    assert any(p in "45 dk kesinti i̇k talebiyle kapatıldı" for p in _HANDOFFS)
    assert any(p in "sorulacaklar" for p in _HANDOFFS)
    assert not any(
        p in "i̇ki yazımın aynı kişi olduğu varsayıldı; yanlışsa iki kişinin "
             "saatleri birleşmiş olur" for p in _HANDOFFS)
    assert not any(p in "süre sayıldı — kontrol edilmeli" for p in _HANDOFFS), \
        "an impersonal 'should be checked' names nobody and stays"


# --- the window must say where it put things --------------------------------

def test_the_result_panel_names_both_output_paths(tmp_path, settings, monkeypatch):
    """"Veri dosyası oluşturuldu" with no path is not actionable.

    The user could not find the snapshot, and the report path was not shown at all.
    Both are now printed in full, resolved.
    """
    import tkinter as tk
    from mesai import gui
    from mesai.gui import rapor, widgets

    # Through the same helper the window tests use: it decides once whether a display
    # exists and then lets a failure be a failure, instead of reading every transient
    # TclError as "headless" and quietly skipping. This test used to lose itself that
    # way roughly once in three full runs. It also parks the window off-screen.
    from tests.test_gui import _tk_root

    root = _tk_root(tk)
    try:
        app = gui.App(root, config_dir=tmp_path, roster_dir=tmp_path, base=tmp_path)
        report = tmp_path / "out" / "mesai-raporu-2026-07.xlsx"
        data = tmp_path / "veri" / "gonderim-2026-07.json"
        app.report._render(rapor.Result(
            True, "Temmuz 2026 raporu yazıldı", ("Toplam: 1:00",), widgets.OK,
            output=report, snapshot=data))
        shown = app.report.result.get("1.0", "end")
    finally:
        root.destroy()

    assert str(report.resolve()) in shown, "report path missing"
    assert str(data.resolve()) in shown, "snapshot path missing"
    assert "RAPOR DOSYASI" in shown and "VERİ DOSYASI" in shown


# --- the labels are filter keys now (ADR-027) -------------------------------

def test_no_two_kinds_share_a_label():
    """The people screen filters on the label, so a duplicate would merge two groups."""
    labels = [label for label, _s, _e, _g in DESCRIPTIONS.values()]
    assert len(labels) == len(set(labels))


def test_every_label_is_short_enough_to_scan_in_a_dropdown():
    """They used to be sentences. A dropdown of sentences cannot be read at a glance."""
    for kind, (label, _severity, _explanation, _group) in DESCRIPTIONS.items():
        assert len(label) <= 38, f"{kind}: {label!r}"


def test_every_kind_explains_itself():
    """The short label drops the meaning; the explanation has to carry it."""
    for kind, (label, _severity, explanation, _group) in DESCRIPTIONS.items():
        assert explanation, f"{kind} has no explanation"
        assert explanation != label, kind


def test_the_two_ambiguous_pairs_stay_distinct():
    """Both pairs caused a real misreading, and both are filter targets.

    "sadece giriş" reads equally as "only the entry exists" and "only the entry is
    missing" — opposite people. And one reading under 5 minutes is a different question
    from a whole day under 2 hours; they shared the words "Süre çok kısa".
    """
    label = {kind: value[0] for kind, value in DESCRIPTIONS.items()}
    assert label[AnomalyKind.MISSING_ENTRY] == "Giriş yok"
    assert label[AnomalyKind.MISSING_EXIT] == "Çıkış yok"
    assert label[AnomalyKind.SHORT_DAY] == "Günlük süre çok kısa (<2 saat)"


def test_the_remote_pair_names_the_kind_that_actually_fires():
    """Measured on May 2026: REMOTE_REPLACED_NOMINAL 35 days, REMOTE_OVERLAP 0.

    ADR-018 removes the system's default day, so nothing is left to overlap with. The
    plain name therefore belongs to the kind the shipped config produces — it was on
    the unreachable one, which would have made the filter look empty.
    """
    label = {kind: value[0] for kind, value in DESCRIPTIONS.items()}
    assert label[AnomalyKind.REMOTE_REPLACED_NOMINAL] == "Uzaktan + sistem kaydı"
    assert label[AnomalyKind.REMOTE_OVERLAP_REAL] == "Uzaktan + kart kaydı"


def test_the_worklist_carries_a_month_level_notes_own_figures(tmp_path, settings):
    """`Ay büyük ölçüde boş` knows the share it could not account for. Show it.

    The share is per person and `Açıklama` is per note, so before the `Ayrıntı` column
    the only sheet carrying it was the row-per-record audit trail — which is not the
    sheet anybody takes to a meeting. One record, its own words.
    """
    sparse = Anomaly(
        kind=AnomalyKind.NO_ATTENDANCE_DATA, source="izin", source_row=0, key=KEY,
        raw_name="AYŞE DENEME",
        detail="İzin kaydı var, hiçbir kart kaydı yok. Bu kişinin ayı eksik görünüyor")
    collector = Collector()
    collector.add(sparse)

    book = openpyxl.load_workbook(
        _build(tmp_path, settings, anomalies=collector), read_only=True)
    rows = list(book["İnceleme Listesi"].iter_rows(values_only=True))
    header = [c for c in rows[3] if c is not None]
    assert header[9] == "Ayrıntı"

    row = next(r for r in rows[4:] if r and r[4] == "Kart bilgisi yok")
    assert row[9] == sparse.detail


def test_the_worklist_leaves_the_detail_empty_when_several_days_disagree(tmp_path,
                                                                        settings):
    """Printing one day's sentence beside a count of five would misdescribe four days."""
    collector = Collector()
    for day, stamp in ((3, "07:41"), (9, "08:02")):
        collector.add(Anomaly(
            kind=AnomalyKind.MISSING_EXIT, source="macunkoy", source_row=day, key=KEY,
            raw_name="AYŞE DENEME", date=date(2026, 5, day), raw_entry=stamp,
            detail=f"giriş {stamp}, çıkış yok"))

    book = openpyxl.load_workbook(
        _build(tmp_path, settings, anomalies=collector), read_only=True)
    rows = list(book["İnceleme Listesi"].iter_rows(values_only=True))
    row = next(r for r in rows[4:] if r and r[4] == "Çıkış yok")

    assert row[6] == 2
    assert not row[9]


def test_the_worklist_prints_the_explanation_beside_the_keyword(tmp_path, settings):
    book = openpyxl.load_workbook(_build(tmp_path, settings), read_only=True)
    rows = list(book["İnceleme Listesi"].iter_rows(values_only=True))
    header = [c for c in rows[3] if c is not None]

    assert header[4] == "Sorun" and header[5] == "Açıklama"
    explanations = {row[4]: row[5] for row in rows[4:] if row and row[4]}
    assert explanations["Çıkış yok"] == "Giriş basılmış, çıkış kaydı yok"


# --- the Etki column says what happened to the RECORD (ADR-055) --------------

def _empty_record(day, row=7):
    return Anomaly(
        kind=AnomalyKind.EMPTY_RECORD, source="macunkoy", source_row=row, key=KEY,
        raw_name="AYŞE DENEME", date=day, detail="giriş de çıkış da boş")


def test_etki_does_not_claim_a_day_was_lost_when_it_counted(tmp_path, settings):
    """The bug: `Bu gün 0 saat sayıldı` on a day that counted 8 hours.

    Severity is a property of the record — the code has always said so — while the
    sentence was about the day. On real data the two disagreed on 52 / 99 / 90 rows over
    May-July 2026, every one of them a Teknopark employee whose Macunköy row was blank
    and whose Teknopark record covered the whole day.
    """
    collector = Collector()
    collector.add(_empty_record(DAY))

    book = openpyxl.load_workbook(
        _build(tmp_path, settings, anomalies=collector), read_only=True)
    rows = list(book["Şüpheli Kayıtlar"].iter_rows(values_only=True))
    header = [c for c in rows[3] if c is not None]
    etki = rows[4][header.index("Etki")]

    # _workday() is DAY and counts 9:39, so the day was NOT lost
    assert "0 saat" not in etki, etki
    assert "başka kayıttan" in etki


def test_etki_still_says_zero_when_the_day_really_counted_nothing(tmp_path, settings):
    """The other half. A rule that never fires in the original direction is not a fix."""
    collector = Collector()
    collector.add(_empty_record(date(2026, 5, 19)))          # no workday on that date

    book = openpyxl.load_workbook(
        _build(tmp_path, settings, anomalies=collector), read_only=True)
    rows = list(book["Şüpheli Kayıtlar"].iter_rows(values_only=True))
    header = [c for c in rows[3] if c is not None]

    assert rows[4][header.index("Etki")] == "Bu gün 0 saat sayıldı"


def test_a_grouped_row_splits_instead_of_picking_one_verdict(tmp_path, settings):
    """One counted, one lost — saying either alone misdescribes the other."""
    collector = Collector()
    collector.add(_empty_record(DAY, row=7))               # counted
    collector.add(_empty_record(date(2026, 5, 19), row=8))    # lost

    book = openpyxl.load_workbook(
        _build(tmp_path, settings, anomalies=collector), read_only=True)
    rows = list(book["İnceleme Listesi"].iter_rows(values_only=True))
    header = [c for c in rows[3] if c is not None]
    row = next(r for r in rows[4:] if r and r[4] == "Hem giriş hem çıkış yok")
    etki = row[header.index("Etki")]

    assert row[header.index("Gün Sayısı")] == 2
    assert "1 gün 0 saat sayıldı" in etki and "1 gün başka kayıttan" in etki


def test_a_refused_reading_is_not_reported_as_no_record(tmp_path, settings):
    """The day the operator found: an exit stamped at 19:56, and `Günlük Detay` said
    `kayıt yok` while `Şüpheli Kayıtlar` said `Giriş yok` for the same day (ADR-067).

    A one-sided punch yields no interval and therefore no `WorkDay`; reading "nothing
    happened" off that absence is what produced two contradictory statements about one
    day. The row now names the file and the note, and shows the stamp.
    """
    collector = Collector()
    collector.add(Anomaly(
        kind=AnomalyKind.MISSING_ENTRY, source="macunkoy", source_row=9, key=KEY,
        raw_name="AYŞE DENEME", date=date(2026, 5, 6),
        raw_exit="06.05.2026 19:56:17"))

    book = openpyxl.load_workbook(
        _build(tmp_path, settings, anomalies=collector), read_only=True)
    rows = list(book["Günlük Detay"].iter_rows(values_only=True))
    header = [c for c in rows[3] if c is not None]
    row = next(r for r in rows[4:] if r and r[1] == "06.05.2026")

    assert row[header.index("Kaynak")] == "Macunköy"
    assert row[header.index("Etiket")] == "Giriş yok"
    assert row[header.index("Son Çıkış")] == "19:56", "ham damga HH:MM olarak"
    assert not row[header.index("Çalışma Süresi")], "sayılan süre yok"


def test_the_control_sheet_counts_roster_people_the_period_never_mentions(
        tmp_path, settings):
    """The group that appeared on no list, in no count and in no note — ADR-071.

    A roster entry with no badge record and no leave row gets no `Employee` and so no
    row (ADR-011), which is right: there is nothing to report about them. What was
    wrong is that `5. Kapsam` counted only the people who *did* reach the report, so
    nothing anywhere said how many were missing from it. That made them the one group
    the manual check on `Kart bilgisi yok` could not reach — 21 / 27 / 14 people over
    May-July 2026, 16 / 22 / 13 of them at the same site that produces almost every
    `Kart bilgisi yok`.

    Names and the facility split are printed because finding these people by hand is
    the entire point of the line, and the wording says the ambiguity out loud: the
    roster has no hire date, so a late joiner and a missing record look identical
    (ROADMAP Q18).
    """
    path = tmp_path / "kapsam.xlsx"
    workbook.build(
        path=path, period="2026-07", summaries=[_summary()], workdays=[_workday()],
        employees={KEY: _employee()}, leave=[], anomalies=Collector(),
        stats=RunStats(files={"teknopark": "test.xlsx"}, roster_size=181,
                       roster_only=(("KEREM ÖRNEK", "MACUNKÖY TESİSİ"),
                                    ("ZEYNEP DENEME", "DEICO TESİS"))),
        settings=settings, generated_at=datetime(2026, 8, 26, 12, 0),
    )

    rows = list(openpyxl.load_workbook(path, read_only=True)["Kontrol"].iter_rows(
        values_only=True))
    control = "\n".join(" ".join(str(c) for c in r if c is not None) for r in rows)

    assert "Personel listesinde olup bu ayda hiç kaydı olmayan 2" in control
    assert "181 kişiden" in control, "the total it is measured against"
    # Grouped by facility, through the label table and never the roster's raw wording
    # (ADR-026) — `Macunköy` and `Teknopark` are what every other sheet calls them.
    assert "Macunköy 1" in control and "Teknopark 1" in control
    assert "KEREM ÖRNEK" in control and "ZEYNEP DENEME" in control
    # It must not read as an absence of work. No hours, no day count, no verdict.
    assert "0:00" not in control.split("5. Kapsam")[1].split("6.")[0]


def test_the_coverage_line_is_absent_when_the_roster_covers_everybody(
        tmp_path, settings):
    """A zero here is not news, and a line that is always present stops being read.

    The counts above it are unconditional because they are always the answer to
    something; this one is only ever a list of people to chase.
    """
    path = tmp_path / "tam-kapsam.xlsx"
    workbook.build(
        path=path, period="2026-07", summaries=[_summary()], workdays=[_workday()],
        employees={KEY: _employee()}, leave=[], anomalies=Collector(),
        stats=RunStats(files={"teknopark": "test.xlsx"}, roster_size=181),
        settings=settings, generated_at=datetime(2026, 8, 26, 12, 0),
    )

    control = "\n".join(
        " ".join(str(c) for c in r if c is not None)
        for r in openpyxl.load_workbook(path, read_only=True)["Kontrol"].iter_rows(
            values_only=True))
    assert "hiç kaydı olmayan" not in control
