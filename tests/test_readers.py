"""Reader tests against synthetic workbooks.

Fixtures use INVENTED names only — never a real employee. See AGENTS.md §2.3.
"""

from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from mesai.anomalies import AnomalyKind
from mesai.readers import LayoutError, izin, macunkoy, roster, teknopark
from tests.builders import (
    IZIN_HEADERS, MAC_HEADERS, ROSTER_HEADERS, izin_row, mac_row, roster_row,
    write_izin, write_macunkoy, write_roster, write_teknopark,
)

# --------------------------------------------------------------------------
# Roster
# --------------------------------------------------------------------------

def roster_row(login, contact, given, surname, email, facility="DEICO TESİS"):
    return [login, contact, given, surname, None, "Çalışan", email, "DEICO",
            "TEST EKİBİ", facility, "TEST GÖREVİ", "Profil", "Files/", None]


def test_roster_key_ignores_middle_names(tmp_path):
    path = write_roster(tmp_path / "SYST03.xlsx", [
        roster_row("ADENEME", "000000001111", "AYŞE", "DENEME", "a@x.com"),
    ])
    entries, duplicates = roster.read(path)
    assert entries[("AYSE", "DENEME")].email == "a@x.com"
    assert duplicates == []


def test_roster_deduplicates_the_same_person_twice(tmp_path):
    """A surname change leaves the old login open — same contact, same e-mail."""
    path = write_roster(tmp_path / "SYST03.xlsx", [
        roster_row("ADENEME", "000000008803", "AYŞE", "DENEME", "a@x.com"),
        roster_row("AYENI", "000000008803", "AYŞE", "DENEME", "a@x.com"),
    ])
    entries, duplicates = roster.read(path)
    assert len(entries) == 1
    assert len(duplicates) == 1
    assert "ADENEME" in duplicates[0] and "AYENI" in duplicates[0]


def test_roster_raises_on_two_different_people(tmp_path):
    """A real collision must fail the run, not silently merge payroll hours."""
    path = write_roster(tmp_path / "SYST03.xlsx", [
        roster_row("ADENEME1", "000000001111", "AYŞE", "DENEME", "a1@x.com"),
        roster_row("ADENEME2", "000000002222", "AYŞE", "DENEME", "a2@x.com"),
    ])
    with pytest.raises(LayoutError, match="FARKLI kişi"):
        roster.read(path)


def test_roster_requires_email_column(tmp_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TEMPIASUSERS"
    sheet.append([h for h in ROSTER_HEADERS if h != "E-posta"])
    path = tmp_path / "calisan_listesi.xlsx"
    workbook.save(path)
    with pytest.raises(LayoutError, match="E-posta"):
        roster.read(path)


def test_roster_sheet_is_found_by_columns_not_by_name(tmp_path):
    """The exporter renames things freely — SYST03_TEMPIASUSERS.xlsx became
    calisan_listesi.xlsx — so the columns are the contract, not the sheet name."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Bambaska Bir Ad"
    sheet.append(ROSTER_HEADERS)
    sheet.append(roster_row("ADENEME", "000000001111", "AYŞE", "DENEME", "a@x.com"))
    path = tmp_path / "calisan_listesi.xlsx"
    workbook.save(path)

    entries, _ = roster.read(path)
    assert entries[("AYSE", "DENEME")].email == "a@x.com"


def test_roster_ignores_the_partial_second_sheet(tmp_path):
    """The real workbook also holds a `Sayfa1` with only name + e-mail. Picking it
    would silently lose facility, department and job title."""
    workbook = openpyxl.Workbook()
    partial = workbook.active
    partial.title = "Sayfa1"
    partial.append(["isim", "E-posta"])
    partial.append(["AYŞE DENEME", "a@x.com"])

    full = workbook.create_sheet("TEMPIASUSERS")
    full.append(ROSTER_HEADERS)
    full.append(roster_row("ADENEME", "000000001111", "AYŞE", "DENEME", "a@x.com",
                           facility="MACUNKÖY TESİSİ"))
    path = tmp_path / "calisan_listesi.xlsx"
    workbook.save(path)

    entries, _ = roster.read(path)
    assert entries[("AYSE", "DENEME")].facility == "MACUNKÖY TESİSİ"


def test_roster_reads_columns_in_a_different_order(tmp_path):
    """Columns are located by header text, so a reordered export still works."""
    reordered = ["Soyad", "İsim", "E-posta", "Görev", "Tesis", "Bölüm",
                 "Kontak No", "Kullanıcı"]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TEMPIASUSERS"
    sheet.append(reordered)
    sheet.append(["DENEME", "AYŞE", "a@x.com", "TEST GÖREVİ", "DEICO TESİS",
                  "TEST EKİBİ", "000000001111", "ADENEME"])
    path = tmp_path / "calisan_listesi.xlsx"
    workbook.save(path)

    entries, _ = roster.read(path)
    entry = entries[("AYSE", "DENEME")]
    assert entry.email == "a@x.com"
    assert entry.facility == "DEICO TESİS"
    assert entry.job_title == "TEST GÖREVİ"


def test_roster_raises_when_no_sheet_has_the_columns(tmp_path):
    workbook = openpyxl.Workbook()
    workbook.active.title = "Başka"
    workbook.active.append(["Bir", "İki"])
    path = tmp_path / "calisan_listesi.xlsx"
    workbook.save(path)
    with pytest.raises(LayoutError, match="beklenen kolonları taşıyan sayfa"):
        roster.read(path)


# --------------------------------------------------------------------------
# Macunköy
# --------------------------------------------------------------------------

def test_macunkoy_reads_rows_and_drops_card_numbers(settings, tmp_path):
    path = write_macunkoy(tmp_path / "Macunköy.xlsx", [
        mac_row("AYŞE", "DENEME", "8802", "2026-05-21", "08:00:00", "17:00:00", "09:00"),
        mac_row("VELİ", "ÖRNEK", "SN999888", "2026-05-21", "08:00:00", "17:00:00", "09:00"),
    ])
    records, anomalies, rows_read, excluded = macunkoy.read(path, settings)

    assert rows_read == 2
    assert excluded == 0
    assert records[0].badge_id == "8802"
    assert records[1].badge_id is None, "SN card numbers must never reach the report"


def test_macunkoy_excludes_shared_badges(settings, tmp_path):
    path = write_macunkoy(tmp_path / "Macunköy.xlsx", [
        mac_row("ZİYARETÇİ35", "ZİYARETÇİ35", "SN1", "2026-05-21", "08:00:00", "17:00:00"),
        mac_row("STJ20", "STJ20", "SN2", "2026-05-21", "08:00:00", "17:00:00"),
        mac_row("AYŞE", "DENEME", "8802", "2026-05-21", "08:00:00", "17:00:00"),
    ])
    records, _, rows_read, excluded = macunkoy.read(path, settings)
    assert rows_read == 3
    assert excluded == 2
    assert len(records) == 1


def test_macunkoy_passes_bad_data_through_untouched(settings, tmp_path):
    """The reader is faithful: negative durations and missing punches survive."""
    path = write_macunkoy(tmp_path / "Macunköy.xlsx", [
        mac_row("AYŞE", "DENEME", "1", "2026-05-28", "23:59:42", "08:07:06", "-15:-52"),
        mac_row("VELİ", "ÖRNEK", "2", "2026-05-28", "08:00:00", None, None),
    ])
    records, anomalies, _, _ = macunkoy.read(path, settings)
    assert len(records) == 2
    assert records[0].exit < records[0].entry, "not corrected here — rules/ does that"
    assert records[1].exit is None
    assert anomalies == []


def test_macunkoy_raises_on_changed_layout(settings, tmp_path):
    workbook = openpyxl.Workbook()
    workbook.active.append(["Isim", "Tarih"])
    path = tmp_path / "Macunköy.xlsx"
    workbook.save(path)
    with pytest.raises(LayoutError):
        macunkoy.read(path, settings)


# --------------------------------------------------------------------------
# Teknopark — the block layout that caused real data loss
# --------------------------------------------------------------------------

def test_teknopark_parses_string_timestamps(settings, tmp_path):
    """A naive isinstance(v, datetime) check yields zero hours for everyone."""
    path = write_teknopark(tmp_path / "Teknopark.xlsx", [{
        "row": 6, "col": 1, "name": "AYŞE DENEME",
        "rows": [("2026-05-04", "04.05.2026 07:35", "04.05.2026 18:51", "11:16")],
        "total": "11:16",
    }])
    records, anomalies, rows_read = teknopark.read(path, settings)

    assert rows_read == 1
    assert records[0].entry == datetime(2026, 5, 4, 7, 35)
    assert records[0].exit == datetime(2026, 5, 4, 18, 51)
    assert anomalies == []


def test_teknopark_survives_irregular_header_offset(settings, tmp_path):
    """Header sits +1 in 61 real blocks, +2 in 9 and +3 in 40."""
    path = write_teknopark(tmp_path / "Teknopark.xlsx", [
        {"row": 6, "col": 1, "name": "AYŞE DENEME", "header_offset": 1,
         "rows": [("2026-05-04", "04.05.2026 08:00", "04.05.2026 17:00", "09:00")],
         "total": "09:00"},
        {"row": 20, "col": 1, "name": "VELİ ÖRNEK", "header_offset": 3,
         "rows": [("2026-05-05", "05.05.2026 08:00", "05.05.2026 16:00", "08:00")],
         "total": "08:00"},
    ])
    records, anomalies, _ = teknopark.read(path, settings)

    assert len(records) == 2
    assert {r.raw_name for r in records} == {"AYŞE DENEME", "VELİ ÖRNEK"}
    assert anomalies == []


def test_teknopark_does_not_stop_at_a_blank_row(settings, tmp_path):
    """The regression that lost 838 of 1 607 rows: blocks are blank-row separated."""
    path = write_teknopark(tmp_path / "Teknopark.xlsx", [{
        "row": 6, "col": 1, "name": "AYŞE DENEME", "header_offset": 2,
        "blank_after": True,
        "rows": [
            ("2026-05-04", "04.05.2026 08:00", "04.05.2026 17:00", "09:00"),
            ("2026-05-05", "05.05.2026 08:00", "05.05.2026 17:00", "09:00"),
            ("2026-05-06", "06.05.2026 08:00", "06.05.2026 17:00", "09:00"),
        ],
        "total": "27:00",
    }])
    records, anomalies, rows_read = teknopark.read(path, settings)

    assert rows_read == 3, "blank spacer rows must not terminate the block"
    assert len(records) == 3
    assert anomalies == []


def test_teknopark_reads_both_column_streams(settings, tmp_path):
    """Blocks in column A and column H are independent and may be misaligned."""
    path = write_teknopark(tmp_path / "Teknopark.xlsx", [
        {"row": 6, "col": 1, "name": "AYŞE DENEME", "header_offset": 3,
         "rows": [("2026-05-04", "04.05.2026 08:00", "04.05.2026 17:00", "09:00"),
                  ("2026-05-05", "05.05.2026 08:00", "05.05.2026 17:00", "09:00")],
         "total": "18:00"},
        {"row": 7, "col": 8, "name": "VELİ ÖRNEK", "header_offset": 1,
         "rows": [("2026-05-04", "04.05.2026 09:00", "04.05.2026 18:00", "09:00")],
         "total": "09:00"},
    ])
    records, anomalies, _ = teknopark.read(path, settings)

    by_name = {r.raw_name: [] for r in records}
    for record in records:
        by_name[record.raw_name].append(record)
    assert len(by_name["AYŞE DENEME"]) == 2
    assert len(by_name["VELİ ÖRNEK"]) == 1
    assert anomalies == []


def test_teknopark_reports_a_block_total_mismatch(settings, tmp_path):
    """The parser's own alarm: if rows are lost, the file's total disagrees."""
    path = write_teknopark(tmp_path / "Teknopark.xlsx", [{
        "row": 6, "col": 1, "name": "AYŞE DENEME",
        "rows": [("2026-05-04", "04.05.2026 08:00", "04.05.2026 17:00", "09:00")],
        "total": "18:00",     # claims twice what the rows add up to
    }])
    _, anomalies, _ = teknopark.read(path, settings)
    assert any(a.kind is AnomalyKind.DURATION_MISMATCH
               and "blok toplamı" in a.detail for a in anomalies)


def test_teknopark_raises_when_the_marker_disappears(settings, tmp_path):
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "Bambaşka bir rapor"
    path = tmp_path / "Teknopark.xlsx"
    workbook.save(path)
    with pytest.raises(LayoutError, match="Adı Soyadı"):
        teknopark.read(path, settings)


# --------------------------------------------------------------------------
# İzin
# --------------------------------------------------------------------------

def test_izin_skips_the_subtotal_row(settings, tmp_path):
    """Counting it doubles every person's leave — docs/DATA-SOURCES.md D6."""
    path = write_izin(tmp_path / "IZIN.xlsx", [
        izin_row("8801", "AYŞE DENEME", None, None, None, None, None, 1.5),
        izin_row("8801", "AYŞE DENEME", "Yıllık İzin", "25.05.2026", "07:30",
                 "25.05.2026", "16:30", 1.0),
        izin_row("8801", "AYŞE DENEME", "Yıllık İzin", "26.05.2026", "07:30",
                 "26.05.2026", "12:00", 0.5),
    ])
    leave, remote, _, rows_read, subtotals = izin.read(path, settings)

    assert rows_read == 3
    assert subtotals == 1
    assert len(leave) == 2
    assert sum(r.days for r in leave) == 1.5
    assert remote == []


def test_izin_turns_remote_work_into_a_punch(settings, tmp_path):
    path = write_izin(tmp_path / "IZIN.xlsx", [
        izin_row("1010", "VELİ ÖRNEK", "Uzaktan Çalışma", "06.05.2026", "07:30",
                 "06.05.2026", "16:30", 1.0),
    ])
    leave, remote, _, _, _ = izin.read(path, settings)

    assert len(remote) == 1
    assert remote[0].tag == "uzaktan"
    assert remote[0].entry == datetime(2026, 5, 6, 7, 30)
    assert remote[0].exit == datetime(2026, 5, 6, 16, 30)
    assert len(leave) == 1, "the record still appears in the leave list, typed"


def test_izin_leaves_other_types_as_absence(settings, tmp_path):
    path = write_izin(tmp_path / "IZIN.xlsx", [
        izin_row("1", "A DENEME", "Yıllık İzin", "05.05.2026", "07:30",
                 "05.05.2026", "16:30", 1.0),
        izin_row("2", "B DENEME", "Mazeret", "05.05.2026", "07:30",
                 "05.05.2026", "10:00", 0.28),
    ])
    _, remote, _, _, _ = izin.read(path, settings)
    assert remote == []


def test_izin_splits_a_multi_day_remote_record(settings, tmp_path):
    path = write_izin(tmp_path / "IZIN.xlsx", [
        izin_row("1", "AYŞE DENEME", "Uzaktan Çalışma", "04.05.2026", "07:30",
                 "06.05.2026", "16:30", 3.0),
    ])
    _, remote, anomalies, _, _ = izin.read(path, settings)

    assert len(remote) == 3
    assert {r.date.day for r in remote} == {4, 5, 6}
    assert any(a.kind is AnomalyKind.MULTI_DAY_REMOTE for a in anomalies)


# --- a file the library cannot open at all -----------------------------------
#
# Wrong COLUMNS already produced a clear LayoutError. A file that is not a workbook at
# all reached the user as `BadZipFile: File is not a zip file`, which names a library
# and a container format and helps nobody — and it is exactly what a renamed CSV, a
# half-downloaded file or a corrupt one produces.

def test_a_file_that_is_not_a_workbook_says_so_plainly(tmp_path):
    from mesai.readers.base import LayoutError, open_sheets

    fake = tmp_path / "liste.xlsx"
    fake.write_text("bu bir excel değil", encoding="utf-8")

    with pytest.raises(LayoutError) as caught:
        open_sheets(fake)

    message = str(caught.value)
    assert "liste.xlsx" in message, "name the file"
    assert "bozuk olabilir" in message, "and say what to do"


def test_the_roster_reader_refuses_a_workbook_with_the_wrong_columns(tmp_path):
    """A wrong file must fail loudly rather than produce empty people."""
    import openpyxl
    from mesai.readers import roster
    from mesai.readers.base import LayoutError

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["Tarih", "Giriş", "Çıkış"])
    sheet.append(["01.06.2026", "08:00", "18:00"])
    path = tmp_path / "yanlis.xlsx"
    book.save(path)

    with pytest.raises(LayoutError, match="beklenen kolonları"):
        roster.read(path)


# --- the roster's header: searched for, and synonyms accepted (ADR-080) -------

def _roster_book(path, header, rows, *, title=None, sheet_name="TEMPIASUSERS"):
    """A roster workbook with a header we choose, optionally under a title row."""
    import openpyxl

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = sheet_name
    if title is not None:
        sheet.append([title])
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    book.save(path)
    return path


_ROSTER_ROW = ["adeneme", "1001", "AYŞE", "DENEME", None, "Çalışan",
               "ayse@example.com", "DEICO", "TEST EKİBİ", "DEICO TESİS",
               "TEST GÖREVİ", "Profil", "Files/", None]


def test_the_roster_header_does_not_have_to_be_the_first_row(tmp_path):
    """The Macunköy export gained a title line above its header in July 2026 and its
    reader was fixed to search (ADR-020). This one read row 1 and nothing else, so the
    same change would have failed it — at month end, on a machine where nobody can edit
    code.
    """
    path = _roster_book(tmp_path / "calisan.xlsx", ROSTER_HEADERS, [_ROSTER_ROW],
                        title="PERSONEL LİSTESİ — 27.08.2026 tarihli dökümdür")

    entries, _dupes = roster.read(path)

    assert list(entries) == [("AYSE", "DENEME")]
    assert entries[("AYSE", "DENEME")].email == "ayse@example.com"
    # and the title row must not have become a person
    assert len(entries) == 1


def test_ad_is_accepted_where_isim_is_expected(tmp_path):
    """*"isim yerine ad falan da yazabilir belki"* — and it is the same column.

    Refusing a synonym would be the program being fussy about wording rather than about
    correctness, which is the opposite of the trade this project makes everywhere else.
    """
    header = ["Kullanıcı Adı", "Sicil No", "Ad", "Soyadı", "Açıklama", "Kontak Tipi",
              "E-Posta Adresi", "Firma", "Departman", "Lokasyon", "Ünvan",
              "Profil", "Dosya Yolu", "Sektör Kodu"]
    path = _roster_book(tmp_path / "personel.xlsx", header, [_ROSTER_ROW])

    entries, _dupes = roster.read(path)
    entry = entries[("AYSE", "DENEME")]

    assert entry.email == "ayse@example.com"
    assert entry.department == "TEST EKİBİ"
    assert entry.facility == "DEICO TESİS"
    assert entry.job_title == "TEST GÖREVİ"


def test_case_and_turkish_characters_do_not_break_the_match(tmp_path):
    """Compared folded, so `E-POSTA`, `e-posta` and `Eposta` are one thing.

    This matters more than it looks: the same fold is what stops `İ` and `I` from being
    two different columns, and Turkish casing is where naive code in this project has
    gone wrong before (AGENTS §2.4).
    """
    header = ["KULLANICI", "kontak no", "isim", "SOYAD", "Açıklama", "Kontak Tipi",
              "eposta", "Firma", "BÖLÜM", "tesis", "GÖREV",
              "Profil", "Dosya Yolu", "Sektör Kodu"]
    path = _roster_book(tmp_path / "calisan.xlsx", header, [_ROSTER_ROW])

    entries, _dupes = roster.read(path)
    assert entries[("AYSE", "DENEME")].email == "ayse@example.com"


def test_extra_and_reordered_columns_are_ignored(tmp_path):
    """Only the named columns are read, wherever they are."""
    header = ["E-posta", "YENİ KOLON", "Soyad", "Başka Bir Şey", "İsim",
              "Kontak No", "Kullanıcı", "Tesis"]
    row = ["ayse@example.com", "alakasiz", "DENEME", 42, "AYŞE", "1001", "adeneme",
           "DEICO TESİS"]
    path = _roster_book(tmp_path / "calisan.xlsx", header, [row])

    entries, _dupes = roster.read(path)
    entry = entries[("AYSE", "DENEME")]

    assert entry.email == "ayse@example.com"
    assert entry.facility == "DEICO TESİS"
    assert entry.department is None, "olmayan opsiyonel kolon boş geçmeli"


def test_a_renamed_required_column_stops_the_run_and_says_what_is_accepted(tmp_path):
    """Loud, never silent. A missing `E-posta` means nobody can be written to, and the
    message has to be actionable by whoever exported the file — so it lists the
    spellings rather than only naming the field.
    """
    header = [h if h != "E-posta" else "İletişim" for h in ROSTER_HEADERS]
    path = _roster_book(tmp_path / "calisan.xlsx", header, [_ROSTER_ROW])

    with pytest.raises(LayoutError) as hata:
        roster.read(path)

    metin = str(hata.value)
    assert "E-posta" in metin
    assert "E-mail" in metin, "kabul edilen yazımlar sayılmalı"
    assert "10 satırda" in metin, "nerede arandığı yazılmalı"


def test_the_thin_second_sheet_is_still_not_picked(tmp_path):
    """The real workbook holds a `Sayfa1` with just name and e-mail. Widening the
    accepted spellings must not let it win — it has no login and no contact number, so
    it cannot satisfy the required set.
    """
    import openpyxl

    path = tmp_path / "calisan.xlsx"
    book = openpyxl.Workbook()
    ince = book.active
    ince.title = "Sayfa1"
    ince.append(["Ad", "E-posta"])
    ince.append(["AYŞE DENEME", "ayse@example.com"])
    dogru = book.create_sheet("Personel Dökümü")
    dogru.append(ROSTER_HEADERS)
    dogru.append(_ROSTER_ROW)
    book.save(path)

    entries, _dupes = roster.read(path)

    assert list(entries) == [("AYSE", "DENEME")]
    assert entries[("AYSE", "DENEME")].job_title == "TEST GÖREVİ", \
        "ince sayfa seçilmiş olurdu"


def test_no_two_fields_claim_the_same_spelling():
    """The table is hand-written, so a duplicate is a real possibility — and it happened
    on the first run: `Kullanıcı` and `Kullanici` fold to one thing. The module asserts
    at import; this makes the reason visible in the suite.
    """
    from mesai.normalize import fold

    seen = {}
    for field, spellings in roster.HEADER_ALIASES.items():
        for spelling in spellings:
            key = fold(spelling)
            assert key not in seen, f"{spelling}: {field} ve {seen[key]}"
            seen[key] = field
