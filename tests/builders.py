"""Builders for synthetic source workbooks.

Extracted from `test_readers.py` so the end-to-end test can assemble a whole month of
input without duplicating the layout knowledge. **Synthetic names only** — the file
layouts are real, the people are not (AGENTS.md §2.3).

If a source layout changes, it changes here once and every test that depends on it
follows.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl

# --- roster (IAS registry) --------------------------------------------------

ROSTER_HEADERS = ["Kullanıcı", "Kontak No", "İsim", "Soyad", "Açıklama",
                  "Kontak Tipi", "E-posta", "Firma", "Bölüm", "Tesis", "Görev",
                  "Profil", "Dosya Yolu", "Sektör Kodu"]


def write_roster(path: Path, rows: list[list]) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TEMPIASUSERS"
    sheet.append(ROSTER_HEADERS)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    return path


def roster_row(login, contact, given, surname, email, facility="DEICO TESİS"):
    return [login, contact, given, surname, None, "Çalışan", email, "DEICO",
            "TEST EKİBİ", facility, "TEST GÖREVİ", "Profil", "Files/", None]


# --- Macunköy (flat log) ----------------------------------------------------

MAC_HEADERS = ["Ad", "Soyad", "Personel", "SicilNo", "Birim", None, "Bolum",
               "MesaiTarih", "Giris", "Cikis", "SureSaat"]


def write_macunkoy(path: Path, rows: list[list],
                   headers: list | None = None,
                   title_row: bool = False) -> Path:
    """Write the flat log.

    `headers` and `title_row` exist because July 2026 dropped the `Personel` column and
    gained a title line above the header (DATA-SOURCES.md D10) — a test needs to be able
    to reproduce both shapes.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sayfa1"
    if title_row:
        sheet.append(["İLK GİRİŞ SON ÇIKIŞ PD"])
    sheet.append(headers if headers is not None else MAC_HEADERS)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    return path


def mac_row(given, surname, badge, day, entry, exit_, duration=None):
    return [given, surname, f"{given} {surname}", badge, "DEICO", None, "EKİP",
            datetime.fromisoformat(f"{day} 00:00:00"),
            datetime.fromisoformat(f"{day} {entry}") if entry else None,
            datetime.fromisoformat(f"{day} {exit_}") if exit_ else None,
            duration]


# --- Teknopark (per-person blocks) -----------------------------------------

def write_teknopark(path: Path, blocks: list[dict]) -> Path:
    """Build a block-layout sheet.

    Each block: {row, col, name, header_offset, rows: [(date, entry, exit, dur)],
                 blank_after: bool, total: "HH:MM"}

    Entry and exit are written as STRINGS on purpose — that is what the real export
    does, and a reader that checks `isinstance(v, datetime)` silently reports zero
    hours for everybody (DATA-SOURCES.md T3).
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Page1"
    sheet["A1"] = "DÖNEMSEL AYRINTILI PUANTAJ RAPORU"

    for block in blocks:
        row, col = block["row"], block["col"]
        sheet.cell(row=row, column=col, value="Adı Soyadı:")
        sheet.cell(row=row, column=col + 3, value=block["name"])

        header_row = row + block.get("header_offset", 1)
        sheet.cell(row=header_row, column=col, value="Tarih")
        sheet.cell(row=header_row, column=col + 2, value="Giriş Tarih Saat")
        sheet.cell(row=header_row, column=col + 4, value="Çıkış Tarih Saat")
        sheet.cell(row=header_row, column=col + 5, value="Çalışma Süresi")

        current = header_row + 1
        for day, entry, exit_, duration in block["rows"]:
            sheet.cell(row=current, column=col,
                       value=datetime.fromisoformat(f"{day} 00:00:00"))
            sheet.cell(row=current, column=col + 2, value=entry)   # STRING
            sheet.cell(row=current, column=col + 4, value=exit_)   # STRING
            sheet.cell(row=current, column=col + 5, value=duration)
            current += 2 if block.get("blank_after") else 1

        sheet.cell(row=current, column=col + 1,
                   value="Dönemdeki Toplam Çalışma Süresi")
        sheet.cell(row=current, column=col + 5, value=block["total"])

    workbook.save(path)
    return path


def tek_block(name, day_rows, *, row=3, col=1, total="9:00", header_offset=1):
    return {"row": row, "col": col, "name": name, "rows": day_rows,
            "total": total, "header_offset": header_offset}


def tek_day(day, entry, exit_, duration):
    """One data row. Times are `dd.mm.yyyy HH:MM` strings, as in the real file."""
    stamp = datetime.fromisoformat(f"{day} 00:00:00").strftime("%d.%m.%Y")
    return (day, f"{stamp} {entry}", f"{stamp} {exit_}", duration)


# --- leave (HCM export) -----------------------------------------------------

IZIN_HEADERS = ["Sicil No", "Görünen Ad", "Bölüm Kodu", "Görev", "İzin Tipi",
                "Onay Durumu", "İzin Durumu", "Başlangıç Tarihi", "Başlangıç Saati",
                "Bitiş Tarihi", "Bitiş Saati", "Mesai Kaydet", "Bordro Kodu",
                "Açıklama", "İzin Sebebi", "Kullanılan Gün"]


def write_izin(path: Path, rows: list[list]) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "HCMPERS"
    sheet.append(IZIN_HEADERS)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    return path


def izin_row(badge, name, leave_type, start_d, start_t, end_d, end_t, days):
    return [badge, name, "EKİP", "GÖREV", leave_type, "Onaylandı", "Kullanıldı",
            start_d, start_t, end_d, end_t, "Mesai Kaydetme", "NORM", None, None, days]
