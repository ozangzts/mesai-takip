"""The report workbook — see docs/OUTPUT-SPEC.md.

Sheet modules write cells. They never compute anything: every figure arrives
already calculated.
"""

from __future__ import annotations

import os
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from ..anomalies import (DESCRIPTIONS, IMPACT_TEXT, TAG_TEXT, Anomaly,
                         AnomalyKind,
                         Collector)
from ..config import Settings
from ..models import Employee, LeaveRecord, MonthSummary, NameKey, RunStats, WorkDay
from ..normalize import sort_key
from ..rules.worktime import clock, decimal_hours, hhmm
from . import styles

class ReportLocked(Exception):
    """The output file could not be replaced — typically open in Excel."""


_DAY_NAMES = ("Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz")

# Worst first, expected-behaviour last: the reader should hit the hours-losing
# problems before the informational rows.
_SEVERITY_ORDER = {"excluded": 0, "included": 1, "info": 2}


def _hours_rule_note(settings: Settings) -> str:
    """One sentence stating the active calculation rule, on every hours sheet.

    The rule is a config switch, so the report must say which way it ran rather than
    leave the reader to assume last month's rule still applies.
    """
    if settings.daily_hours == "union":
        measure = ("Süre, gün içindeki giriş-çıkış aralıklarının toplamıdır; "
                   "aralar arasındaki boşluklar sayılmaz")
    else:
        measure = ("Süre, günün ilk girişinden son çıkışına kadar hesaplanır; "
                   "gün içindeki boşluklar düşülmez")
    if settings.brk.deduct:
        brk = f"öğle arası için {settings.brk.minutes} dk kesinti uygulanır"
    else:
        brk = "öğle arası için kesinti UYGULANMAZ"
    return f"HESAP KURALI: {measure}. Ayrıca {brk}."

# How an interval's origin is shown to the reader.
# `kayit-yok` is not a file. The day was derived from the calendar rather than read
# from anywhere, and saying "Uzaktan" or naming a site would be a claim about a record
# that does not exist. ADR-060.
_SOURCE_LABEL = {"macunkoy": "Macunköy", "teknopark": "Teknopark", "izin": "Uzaktan",
                 "kayit-yok": "kayıt yok"}
# How an input file is named on the Kontrol sheet. Still its own table rather than a
# reuse of the one above, for one reason that has not changed: the leave file is
# "İzin", even though the intervals it contributes are labelled "Uzaktan".
#
# These were "Macunköy giriş-çıkış", "Teknopark puantaj" and "İzin (HCM)". The suffixes
# described how each site happens to record attendance, which is a distinction the
# reader never has to make — and next to the file name they were saying it twice, since
# the row already reads "Macunköy giriş-çıkış: Macunköy Mayıs Mesai giriş-çıkış.xlsx".
# The window uses the same three words, so the two now agree.
_FILE_LABEL = {
    "roster": "Personel listesi", "macunkoy": "Macunköy",
    "teknopark": "Teknopark", "izin": "İzin",
}


def build(
    path: Path,
    period: str,
    summaries: list[MonthSummary],
    workdays: list[WorkDay],
    employees: dict[NameKey, Employee],
    leave: list[LeaveRecord],
    anomalies: Collector,
    stats: RunStats,
    settings: Settings,
    generated_at: datetime,
) -> None:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    footer = _footer_lines(period, stats, generated_at)

    # What each person-day actually counted. The `Etki` column needs it: severity is a
    # property of the RECORD ("this record contributed no hours") while the sentence it
    # used to print was about the DAY ("Bu gün 0 saat sayıldı"), and the two disagree on
    # 52 / 99 / 90 rows over May-July 2026 — every one of them a day that counted eight
    # hours or more. See ADR-055.
    measured = {(w.key, w.date): w.gross for w in workdays}

    _sheet_summary(workbook.create_sheet("Aylık Özet"), period, summaries,
                   anomalies, stats, settings, footer)
    _sheet_daily(workbook.create_sheet("Günlük Detay"), workdays, employees,
                 leave, anomalies, settings, footer, period)
    _sheet_worklist(workbook.create_sheet("İnceleme Listesi"), period, anomalies,
                    employees, settings, footer, measured)
    _sheet_anomalies(workbook.create_sheet("Şüpheli Kayıtlar"), anomalies,
                     employees, footer, measured)
    _sheet_leave(workbook.create_sheet("İzin Özeti"), leave, employees,
                 summaries, settings, footer)
    _sheet_control(workbook.create_sheet("Kontrol"), period, stats, summaries,
                   anomalies, settings, generated_at)

    # Write to a temporary file and move it into place, so a crash mid-write cannot
    # leave HR with a half-written workbook that still opens.
    target = path.with_suffix(".tmp.xlsx")
    workbook.save(target)
    workbook.close()
    try:
        os.replace(target, path)
    except PermissionError as exc:
        target.unlink(missing_ok=True)
        raise ReportLocked(
            f"Rapor yazılamadı, dosya kilitli:\n  {path}\n\n"
            "Dosya Excel'de açıksa kapatıp tekrar deneyin."
        ) from exc


# ---------------------------------------------------------------------------
# Sheet 1 — Aylık Özet
# ---------------------------------------------------------------------------

_SUMMARY_HEAD = ["Ad Soyad", "Sicil No", "Departman", "Görev", "Tesis",
                 "Kayıt Kaynağı", "Çalışılan Gün"]
_SUMMARY_HEAD_WIDTHS = [28, 10, 30, 32, 17, 20, 13]
_SUMMARY_TAIL = ["Uzaktan Çalışma (Gün)", "İzin Günü", "Şüpheli Kayıt", "Not"]
# The Not column carries every note the person has, in the same words as the filter
# list. Measured over three months: at most four notes, the longest text 93 characters.
_SUMMARY_TAIL_WIDTHS = [14, 11, 12, 52]


def _hours_columns(settings: Settings) -> tuple[list[str], list[int]]:
    """The hours columns, which depend on whether a break is deducted.

    With the deduction off, gross and net are the same number. Printing both would
    make an HR reader ask which one payroll uses — so one pair of columns is shown,
    named for what it is. See ADR-016.
    """
    if settings.brk.deduct:
        return (["Brüt Süre", "Brüt (Saat)", "Net Süre", "Net (Saat)"],
                [11, 11, 11, 11])
    return ["Çalışma Süresi", "Çalışma (Saat)"], [14, 14]


def _sheet_summary(sheet: Worksheet, period: str, summaries: list[MonthSummary],
                   anomalies: Collector, stats: RunStats, settings: Settings,
                   footer: list[str]) -> None:
    hours_headers, hours_widths = _hours_columns(settings)
    headers = _SUMMARY_HEAD + hours_headers + _SUMMARY_TAIL
    widths = _SUMMARY_HEAD_WIDTHS + hours_widths + _SUMMARY_TAIL_WIDTHS
    span = len(headers)

    # Column positions are derived, never hard-coded: the hours block changes width
    # with the config and off-by-one formatting here would mislabel payroll figures.
    first_hours = len(_SUMMARY_HEAD) + 1
    hours_cols = range(first_hours, first_hours + len(hours_headers))
    decimal_cols = {c for i, c in enumerate(hours_cols) if i % 2 == 1}
    right_cols = set(hours_cols) | {len(_SUMMARY_HEAD)}          # + "Çalışılan Gün"
    day_cols = {first_hours + len(hours_headers), first_hours + len(hours_headers) + 1}
    right_cols |= day_cols | {first_hours + len(hours_headers) + 2}

    styles.write_title(sheet, 1, f"AYLIK ÇALIŞMA ÖZETİ — {_period_label(period)}", span)
    styles.write_banner(
        sheet, 2,
        "DİKKAT: Şüpheli olarak işaretlenen kayıtlar 0 saat sayılmıştır — "
        "'Şüpheli Kayıtlar' sayfasına bakın. Bu rapor bir DOĞRULAMA koşusudur, "
        "bordro için nihai değildir.", span)
    styles.write_banner(sheet, 3, _hours_rule_note(settings), span)
    header_row = 4
    partial = [c for c in stats.coverage.values() if c.is_partial]
    if partial or stats.blank_workdays:
        # The loudest thing on the deliverable sheet. A report built from a mid-month
        # export looks completely normal otherwise — ADR-020.
        parts = ["; ".join(
            f"{_FILE_LABEL.get(c.source, c.source)}: "
            f"{c.trailing_missing[0]:%d.%m} tarihinden sonrası yok "
            f"({c.present}/{c.expected} iş günü)"
            for c in sorted(partial, key=lambda c: c.source))] if partial else []
        if stats.blank_workdays:
            parts.append(
                f"{len(stats.blank_workdays)} iş gününde hiçbir tesiste kayıt yok "
                f"({', '.join(d.strftime('%d.%m') for d in stats.blank_workdays[:8])}"
                + (" ..." if len(stats.blank_workdays) > 8 else "") + ")")
        detail = ". ".join(parts)
        styles.write_banner(
            sheet, 4,
            f"BU RAPOR EKSİK — dönemin tamamı kapsanmıyor. {detail}. "
            "Saatler bordro için KULLANILAMAZ; kaynak dosyalar ya da tatil listesi "
            "gözden geçirilmeli. Ayrıntı: 'Kontrol' sayfası, bölüm 3.", span)
        for column in range(1, span + 1):
            sheet.cell(row=4, column=column).fill = styles.RED_FILL
        header_row = 5
    styles.write_header(sheet, header_row, headers, widths)

    row = header_row + 1
    total_gross = timedelta()
    total_net = timedelta()

    for summary in sorted(summaries, key=lambda s: sort_key(s.employee.display_name)):
        employee = summary.employee
        fill = None
        # Red means "you cannot use this row's hours". Two ways to get there and both
        # earn it: no record at all, and records that could none of them be counted
        # (ADR-067). The second used to be indistinguishable from the first.
        if not summary.has_attendance or not summary.worked_days:
            fill = styles.RED_FILL
        elif summary.anomaly_count:
            fill = styles.AMBER_FILL

        values: list[object] = [
            employee.display_name,
            employee.personnel_no or "",
            employee.department or "",
            employee.job_title or "",
            settings.facility(employee.facility),
            _sources_label(employee.sources),
        ]
        if summary.has_attendance:
            hours: list[object] = [hhmm(summary.gross), decimal_hours(summary.gross)]
            if settings.brk.deduct:
                hours += [hhmm(summary.net), decimal_hours(summary.net)]
            values += [summary.worked_days] + hours
            total_gross += summary.gross
            total_net += summary.net
        else:
            values += [""] * (1 + len(hours_headers))
        values += [
            summary.remote_days or "",
            summary.leave_days or "",
            summary.anomaly_count or "",
            "; ".join(summary.notes),
        ]

        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            if index in right_cols:
                cell.alignment = styles.RIGHT
            if index in decimal_cols:
                cell.number_format = "0.00"
            if index in day_cols:
                cell.number_format = "0.0#"
        styles.style_row(sheet, row, span, fill)
        row += 1

    with_attendance = sum(1 for s in summaries if s.has_attendance)
    sheet.cell(row=row, column=1, value=f"TOPLAM ({with_attendance} kişi)")
    totals: list[object] = [hhmm(total_gross), decimal_hours(total_gross)]
    if settings.brk.deduct:
        totals += [hhmm(total_net), decimal_hours(total_net)]
    for offset, value in enumerate(totals):
        cell = sheet.cell(row=row, column=first_hours + offset, value=value)
        cell.alignment = styles.RIGHT
    for index in range(1, span + 1):
        cell = sheet.cell(row=row, column=index)
        cell.font = styles.TOTAL_FONT
        cell.fill = styles.TOTAL_FILL
        cell.border = styles.BORDER
    for index in decimal_cols:
        sheet.cell(row=row, column=index).number_format = "0.00"

    styles.write_footer(sheet, row + 2, footer, span)


# ---------------------------------------------------------------------------
# Sheet 2 — Günlük Detay
# ---------------------------------------------------------------------------

# The words a day with no record at all carries in `Etiket`. Taken from the note rather
# than written here, because one fact has one wording (ADR-049, ADR-050) — this is the
# same day `Hem giriş hem çıkış yok` is raised for.
TAG_TEXT_EMPTY = DESCRIPTIONS[AnomalyKind.EMPTY_RECORD][0]

_DAILY_HEAD = ["Ad Soyad", "Tarih", "Gün", "İlk Giriş", "Son Çıkış", "Aralık Sayısı"]
_DAILY_HEAD_WIDTHS = [28, 12, 13, 10, 10, 13]
_DAILY_TAIL = ["Kaynak", "Etiket"]
_DAILY_TAIL_WIDTHS = [24, 26]


def _daily_rows(
    workdays: list[WorkDay], employees: dict[NameKey, Employee],
    leave: list[LeaveRecord], settings: Settings, period: str,
) -> list[tuple[NameKey, date, WorkDay | None, str]]:
    """Every person, every day — `(key, date, workday or None, leave type)`.

    The sheet used to hold one row per `WorkDay`, so a day with no usable record simply
    had no row and the reader could not tell "did not come in" from "not in this sheet".
    Now it carries, for every person in the report:

    * every **expected working day** of the period, whether or not anything was recorded,
    * plus any weekend or holiday on which the person **did** work.

    Weekends and holidays with no record are left out: nobody has to account for a day
    they were not expected. A holiday that was worked appears, shaded, because that is the
    day somebody will ask about.

    This is a **display** expansion and nothing else. No `WorkDay` is invented, so the
    reconciliation invariant (Σ per-person == Σ measured person-days) is untouched — the
    added rows carry no hours because there are none to carry.
    """
    year, month = (int(part) for part in period.split("-"))
    expected = settings.calendar.expected_workdays(year, month)

    on_leave: dict[tuple[NameKey, date], str] = {}
    for record in leave:
        if record.start is None:
            continue
        day, last = record.start.date(), (record.end or record.start).date()
        while day <= last:
            on_leave.setdefault((record.key, day), record.leave_type)
            day += timedelta(days=1)

    measured = {(w.key, w.date): w for w in workdays}
    wanted: set[tuple[NameKey, date]] = {(key, day) for key in employees
                                         for day in expected}
    wanted |= set(measured)          # weekend and holiday work keeps its row

    rows = []
    for key, day in wanted:
        rows.append((key, day, measured.get((key, day)),
                     on_leave.get((key, day), "")))
    return sorted(rows, key=lambda r: (
        sort_key(employees[r[0]].display_name if r[0] in employees else ""), r[1]))


def _sheet_daily(sheet: Worksheet, workdays: list[WorkDay],
                 employees: dict[NameKey, Employee], leave: list[LeaveRecord],
                 anomalies: Collector, settings: Settings, footer: list[str],
                 period: str = "") -> None:
    if settings.brk.deduct:
        middle, middle_widths = ["Brüt", "Öğle Kesintisi", "Net"], [10, 14, 10]
    else:
        # Gaps are paid, so showing them is how a reader checks the day themselves:
        # Son Çıkış − İlk Giriş must equal Çalışma Süresi, and the gap column says
        # how much of it was time away from the badge readers.
        middle, middle_widths = ["Çalışma Süresi", "Gün İçi Boşluk"], [15, 15]

    headers = _DAILY_HEAD + middle + _DAILY_TAIL
    widths = _DAILY_HEAD_WIDTHS + middle_widths + _DAILY_TAIL_WIDTHS
    span = len(headers)
    right_cols = set(range(4, len(_DAILY_HEAD) + len(middle) + 1))

    styles.write_title(sheet, 1, "GÜNLÜK DETAY — özetin denetim izi", span)
    styles.write_banner(
        sheet, 2,
        "Her kişinin her iş günü burada — çalıştığı, izinli olduğu ve hiç kaydı "
        "olmayan günler dahil. Hafta sonu ve tatiller yalnızca o gün çalışılmışsa "
        "görünür. 'Aralık Sayısı' 1'den büyükse gün bölünmüş (ara giriş-çıkış) "
        "demektir. Kaynak 'Macunköy → Teknopark' ise ilk giriş bir tesiste, son çıkış "
        "diğerinde; '+' ise iki tesisin kaydı aynı aralıkta birleşmiştir ve çakışan "
        "süre bir kez sayılmıştır; 'İzin' ya da 'kayıt yok' ise o gün çalışma kaydı "
        "yoktur.", span)
    styles.write_banner(sheet, 3, _hours_rule_note(settings), span)
    styles.write_header(sheet, 4, headers, widths)

    # A day with no measurement is not necessarily a day with no record: a one-sided
    # punch yields no interval and therefore no `WorkDay`. Reading "nothing happened"
    # off the absence of a `WorkDay` printed `kayıt yok` on a day whose exit was
    # stamped at 19:56, while `Şüpheli Kayıtlar` said `Giriş yok` for the same day
    # (ADR-067). The anomalies are what that day actually has to say.
    refused: dict[tuple[NameKey, date], list[Anomaly]] = defaultdict(list)
    for anomaly in anomalies.items:
        if anomaly.key is not None and anomaly.date is not None:
            refused[(anomaly.key, anomaly.date)].append(anomaly)

    row = 5
    for key, day, workday, leave_type in _daily_rows(
            workdays, employees, leave, settings, period):
        employee = employees.get(key)
        day_label = settings.calendar.label(day)

        if workday is None:
            kayitlar = [a for a in refused.get((key, day), ())
                        if a.source in _SOURCE_LABEL and a.source != "kayit-yok"]
            # Nothing measured. The times stay empty rather than 0 — there is no
            # reading, and `00:00` would look like one.
            middle_values = ["", "", ""] if settings.brk.deduct else ["", ""]
            if leave_type:
                # `İzin` and nothing more. The HCM's own type went in this column and it
                # spelled out `Doğum İzni (Tam Ödeme)` and `İstirahat (Raporlu)` against
                # a name and a date — which is a person's parental and medical record on
                # the sheet HR circulates, and is not what the sheet is for. `Kaynak`
                # already says the day was leave; `İzin Özeti` is where the breakdown
                # belongs. ADR-070.
                kaynak, etiket, giris, cikis = "İzin", "", "", ""
            elif kayitlar:
                # The person appears on this day; nothing could be counted from it.
                kaynak = " + ".join(sorted(
                    {_SOURCE_LABEL[a.source] for a in kayitlar}))
                etiket = ", ".join(sorted({a.label for a in kayitlar}))
                giris = next((clock(a.raw_entry) for a in kayitlar if a.raw_entry), "")
                cikis = next((clock(a.raw_exit) for a in kayitlar if a.raw_exit), "")
            else:
                kaynak, etiket, giris, cikis = "kayıt yok", TAG_TEXT_EMPTY, "", ""
            values = [
                employee.display_name if employee else "",
                day.strftime("%d.%m.%Y"),
                day_label,
                giris, cikis, "",
                *middle_values,
                kaynak,
                etiket,
            ]
        else:
            if settings.brk.deduct:
                middle_values = [hhmm(workday.gross), hhmm(workday.break_deduction),
                                 hhmm(workday.net)]
            else:
                middle_values = [hhmm(workday.gross), hhmm(workday.gap_total)]
            values = [
                employee.display_name if employee else "",
                day.strftime("%d.%m.%Y"),
                day_label,
                workday.first_entry.strftime("%H:%M") if workday.first_entry else "",
                workday.last_exit.strftime("%H:%M") if workday.last_exit else "",
                len(workday.intervals),
                *middle_values,
                _day_sources_label(workday),
                _tags_label(workday.tags),
            ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            if index in right_cols:
                cell.alignment = styles.RIGHT
        fill = None
        if workday is None:
            # Grey for leave, red for anything else: a day that counted nothing is a day
            # somebody has to account for, whether a reading was refused or none exists.
            fill = styles.GREY_FILL if leave_type else styles.RED_FILL
        elif workday.tags:
            fill = styles.AMBER_FILL
        if settings.calendar.is_holiday(day):
            fill = styles.GREY_FILL
        styles.style_row(sheet, row, span, fill)
        row += 1

    styles.write_footer(sheet, row + 1, footer, span)


# ---------------------------------------------------------------------------
# Sheet 3 — İnceleme Listesi (per-person worklist)
# ---------------------------------------------------------------------------

# `Açıklama` used to be here, a 52-wide sentence repeated on every row carrying that
# note. It was the same text every time — the note's meaning does not vary by person —
# so it bought nothing per row and cost the width the days needed. It is a legend under
# the table now, once per note that actually occurs (ADR-075).
# `Ayrıntı` stays: it carries the RECORD's own words, and only where the row stands for
# a single record, so it does differ row to row.
_WORKLIST_HEADERS = [
    "Ad Soyad", "Sicil No", "Tesis", "Departman", "Sorun",
    "Gün Sayısı", "Günler", "Etki", "Ayrıntı",
]
_WORKLIST_WIDTHS = [28, 10, 14, 30, 26, 11, 52, 24, 58]


def _single_detail(found: list[str]) -> str:
    """The record's own words, but only when the row stands for one record.

    A month-level note is always one record, which is the case this exists for. A note
    spanning several days has several details and gets none: the reader is sent to the
    audit sheet rather than shown one day's sentence as if it covered the rest.
    """
    distinct = {text for text in found if text}
    return distinct.pop() if len(found) == 1 and len(distinct) == 1 else ""


def _sheet_worklist(sheet: Worksheet, period: str, anomalies: Collector,
                    employees: dict[NameKey, Employee], settings: Settings,
                    footer: list[str], measured: dict) -> None:
    """One row per (person, problem), with the exact dates listed.

    The Şüpheli Kayıtlar sheet is the audit trail — one row per defective record,
    216 of them. This sheet is the thing you take to HR or IT and ask about: it
    collapses those rows into a per-person question with the days named.
    """
    span = len(_WORKLIST_HEADERS)
    styles.write_title(sheet, 1,
                       f"İNCELEME LİSTESİ — kişi bazlı eksik kayıt listesi "
                       f"({_period_label(period)})", span)
    styles.write_banner(
        sheet, 2,
        "Bu sayfa incelenmesi gereken kayıtların listesi: her satır bir kişi ve "
        "bir konu, hangi günlerde olduğu yazılı. KIRMIZI: o günler 0 saat sayıldı, "
        "kayıp saat var. SARI: sayıldı ama bakılması iyi olur. GRİ: beklenen durum, "
        "sorun değil — bilgi için listelenmiştir. Satır bazlı denetim izi için "
        "'Şüpheli Kayıtlar' sayfasına bakın.", span)
    styles.write_header(sheet, 4, _WORKLIST_HEADERS, _WORKLIST_WIDTHS)

    # The label is what rows are grouped by, so the explanation is looked up from it
    # rather than carried on every row.
    explanations = {label: explanation
                    for label, _severity, explanation, _group in DESCRIPTIONS.values()}

    # (employee key, problem label) -> dates
    grouped: dict[tuple[NameKey | None, str], list[date]] = defaultdict(list)
    severity: dict[tuple[NameKey | None, str], str] = {}
    names: dict[NameKey | None, str] = {}
    # Every detail the bucket's records wrote, so the column can be filled only when
    # there is ONE of them. Several records mean several different sentences, and
    # picking one of them to print would be a lie about the other days.
    details: dict[tuple[NameKey | None, str], list[str]] = defaultdict(list)

    for anomaly in anomalies.items:
        bucket = (anomaly.key, anomaly.label)
        if anomaly.date is not None:
            grouped[bucket].append(anomaly.date)
        else:
            grouped[bucket]          # touch, so a dateless problem still gets a row
        details[bucket].append(anomaly.detail)
        severity[bucket] = anomaly.severity
        employee = employees.get(anomaly.key) if anomaly.key else None
        names[anomaly.key] = employee.display_name if employee else anomaly.raw_name

    def order(item):
        """By name, in Turkish alphabetical order, like every other sheet.

        It led with severity and then with descending day count, which put the worst
        case at the top — useful if you are triaging, confusing if you are looking
        somebody up, and this is the sheet people look somebody up in. The colour
        already says the severity on every row, and `Gün Sayısı` is sortable in Excel by
        whoever wants that order. `Aylık Özet`, `Günlük Detay` and `İzin Özeti` are all
        by name; a fourth sheet with its own order is one the reader has to relearn.
        """
        (key, label), _dates = item
        return (sort_key(names.get(key, "")), label)

    row = 5
    for (key, label), dates in sorted(grouped.items(), key=order):
        employee = employees.get(key) if key else None
        unique_days = sorted({d.day for d in dates})

        values = [
            names.get(key, ""),
            employee.personnel_no if employee and employee.personnel_no else "",
            settings.facility(employee.facility if employee else ""),
            employee.department if employee else "",
            label,
            len(unique_days) or "",
            _day_list(unique_days, period),
            _group_impact(severity[(key, label)], dates, key, measured),
            _single_detail(details[(key, label)]),
        ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            if index == 6:                       # Gün Sayısı
                cell.alignment = styles.RIGHT
            if index in (7, 9):                  # Günler, Ayrıntı
                cell.alignment = styles.LEFT
        level = severity[(key, label)]
        fill = {"excluded": styles.RED_FILL, "info": styles.GREY_FILL}.get(
            level, styles.AMBER_FILL)
        styles.style_row(sheet, row, span, fill)
        row += 1

    row = _worklist_legend(sheet, row + 1, grouped, explanations, span)
    styles.write_footer(sheet, row + 1, footer, span)


def _worklist_legend(sheet: Worksheet, row: int, grouped: dict,
                     explanations: dict[str, str], span: int) -> int:
    """The note meanings, once each, under the table instead of on every row.

    Only the notes this month actually produced, in the order they appear in
    `anomalies.py` — a legend listing notes nobody has invites the reader to look for
    people who are not there.

    Below the table rather than on `Kontrol`: the reader who needs a meaning is looking
    at the row that has it, and a meaning on another sheet is a meaning nobody reads.
    """
    seen = [label for _key, label in grouped]
    order = [label for label, _s, _e, _g in DESCRIPTIONS.values()]
    labels = [label for label in dict.fromkeys(order) if label in set(seen)]
    if not labels:
        return row

    heading = sheet.cell(row=row, column=1, value="SORUNLARIN ANLAMI")
    heading.font = styles.HEADER_FONT
    heading.fill = styles.HEADER_FILL
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    row += 1
    for label in labels:
        sheet.cell(row=row, column=1, value=label)
        cell = sheet.cell(row=row, column=2, value=explanations.get(label, ""))
        cell.alignment = styles.LEFT
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=span)
        row += 1
    return row


# Plural wording for this sheet, which groups several days into one row. Only the
# excluded text differs from `anomalies._IMPACT`; the rest defers to it so a new
# severity cannot be added there and silently KeyError here.
_IMPACT_TEXT_PLURAL = {"excluded": "Bu günler 0 saat sayıldı"}


def _impact_text(severity: str) -> str:
    return _IMPACT_TEXT_PLURAL.get(severity, IMPACT_TEXT[severity])


def _record_impact(anomaly: Anomaly, measured: dict) -> str:
    """What happened, for ONE record — and it is not always what the day did.

    `excluded` means this record contributed no hours. It used to print "Bu gün 0 saat
    sayıldı", which is a claim about the day, and on 52 / 99 / 90 rows over May-July 2026
    that claim was false: the day counted eight hours or more from another record. The
    reader was told a day was lost when nothing was.

    Almost all of those are one situation. The person is Teknopark staff who called at
    the Macunköy site; the Macunköy row is blank or half-written, and their Teknopark
    record covers the whole day. Measured on June 2026: 99 of 99.
    """
    if anomaly.severity != "excluded":
        return IMPACT_TEXT[anomaly.severity]
    if anomaly.date is None or anomaly.key is None:
        return IMPACT_TEXT["excluded"]
    gross = measured.get((anomaly.key, anomaly.date))
    if not gross:
        return IMPACT_TEXT["excluded"]
    return f"Bu kayıt sayılmadı; gün başka kayıttan {hhmm(gross)} sayıldı"


def _group_impact(severity: str, dates: list, key, measured: dict) -> str:
    """The same question for a row that stands for several days.

    Split rather than picked: a row reading `15 gün` beside one verdict would describe
    fourteen of them wrongly, which is the mistake `Ayrıntı` avoids in the same sheet.
    """
    if severity != "excluded" or not dates:
        return _impact_text(severity)
    lost = sum(1 for d in dates if not measured.get((key, d)))
    other = len(dates) - lost
    if other == 0:
        return _impact_text("excluded")
    if lost == 0:
        return "Bu kayıtlar sayılmadı; günler başka kayıttan sayıldı"
    return f"{lost} gün 0 saat sayıldı; {other} gün başka kayıttan sayıldı"


_MONTHS_LOWER = ("Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", "Temmuz",
                 "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık")


def _day_list(days: list[int], period: str) -> str:
    """`4, 5, 11, 12 Mayıs` — compact enough to read, precise enough to ask about."""
    if not days:
        return "tüm ay"
    month = _MONTHS_LOWER[int(period.split("-")[1]) - 1]
    return ", ".join(str(d) for d in days) + f" {month}"


# ---------------------------------------------------------------------------
# Sheet 4 — Şüpheli Kayıtlar
# ---------------------------------------------------------------------------

_ANOMALY_HEADERS = [
    "Ad Soyad", "Tarih", "Sorun", "Kaynak Dosya", "Kaynak Satır",
    "Ham Giriş", "Ham Çıkış", "Etki", "Açıklama",
]
_ANOMALY_WIDTHS = [28, 12, 34, 14, 13, 22, 22, 24, 52]


def _sheet_anomalies(sheet: Worksheet, anomalies: Collector,
                     employees: dict[NameKey, Employee], footer: list[str],
                     measured: dict) -> None:
    span = len(_ANOMALY_HEADERS)
    styles.write_title(sheet, 1, "ŞÜPHELİ KAYITLAR", span)
    styles.write_banner(
        sheet, 2,
        "Kırmızı satırlar toplama DAHİL EDİLMEDİ (o gün 0 saat). Sarı satırlar "
        "toplama dahil edildi ama kontrol edilmeli. Gri satırlar beklenen "
        "durumdur, sorun değil — denetim izi tam olsun diye listelenir. "
        "'Kaynak Satır' orijinal dosyadaki satır numarasıdır — açıp "
        "bakabilirsiniz.", span)
    styles.write_header(sheet, 4, _ANOMALY_HEADERS, _ANOMALY_WIDTHS)

    row = 5
    ordered = sorted(
        anomalies.items,
        key=lambda a: (
            _SEVERITY_ORDER[a.severity],
            sort_key(employees[a.key].display_name
                     if a.key in employees else a.raw_name),
            a.date or date.min,
        ),
    )
    for anomaly in ordered:
        employee = employees.get(anomaly.key) if anomaly.key else None
        name = employee.display_name if employee else anomaly.raw_name
        values = [
            name,
            anomaly.date.strftime("%d.%m.%Y") if anomaly.date else "",
            anomaly.label,
            _SOURCE_LABEL.get(anomaly.source, anomaly.source),
            anomaly.source_row,
            anomaly.raw_entry,
            anomaly.raw_exit,
            _record_impact(anomaly, measured),
            anomaly.detail,
        ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            if index == 5:
                cell.alignment = styles.RIGHT
        fill = {"excluded": styles.RED_FILL, "info": styles.GREY_FILL}.get(
            anomaly.severity, styles.AMBER_FILL)
        styles.style_row(sheet, row, span, fill)
        row += 1

    styles.write_footer(sheet, row + 1, footer, span)


# ---------------------------------------------------------------------------
# Sheet 5 — İzin Özeti
# ---------------------------------------------------------------------------

_LEAVE_TYPES = [
    "Yıllık İzin", "Mazeret", "Uzaktan Çalışma", "İstirahat (Raporlu)",
    "Eğitim İzni", "Doğum Günü İzni",
]
_LEAVE_HEADERS = (["Ad Soyad", "Sicil No"] + _LEAVE_TYPES
                  + ["Diğer", "Toplam Gün", "Not"])
_LEAVE_WIDTHS = [28, 10] + [15] * len(_LEAVE_TYPES) + [10, 12, 40]


def _sheet_leave(sheet: Worksheet, leave: list[LeaveRecord],
                 employees: dict[NameKey, Employee], summaries: list[MonthSummary],
                 settings: Settings, footer: list[str]) -> None:
    span = len(_LEAVE_HEADERS)
    styles.write_title(sheet, 1, "İZİN ÖZETİ", span)
    styles.write_banner(
        sheet, 2,
        "'Uzaktan Çalışma' izin değil, ÇALIŞMA olarak sayılmıştır — Aylık Özet'teki "
        "saatlerin içindedir. İzin dosyasındaki serbest metin gerekçeler (sağlık, "
        "kişisel) kişisel veri olduğu için bu rapora yazılmamıştır.", span)
    styles.write_header(sheet, 4, _LEAVE_HEADERS, _LEAVE_WIDTHS)

    by_person: dict[NameKey, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for record in leave:
        by_person[record.key][record.leave_type] += record.days

    has_attendance = {s.employee.key for s in summaries if s.has_attendance}

    row = 5
    for key, types in sorted(
        by_person.items(),
        key=lambda item: sort_key(employees[item[0]].display_name
                                  if item[0] in employees else ""),
    ):
        employee = employees.get(key)
        other = sum(days for name, days in types.items() if name not in _LEAVE_TYPES)
        total = sum(types.values())
        note = "" if key in has_attendance else "Kart bilgisi yok"

        values: list[object] = [
            employee.display_name if employee else "",
            employee.personnel_no if employee and employee.personnel_no else "",
        ]
        values += [round(types.get(name, 0.0), 2) or "" for name in _LEAVE_TYPES]
        values += [round(other, 2) or "", round(total, 2), note]

        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            if 3 <= index <= span - 1:
                cell.alignment = styles.RIGHT
                cell.number_format = "0.0#"
        styles.style_row(sheet, row, span,
                         styles.RED_FILL if note else None)
        row += 1

    styles.write_footer(sheet, row + 1, footer, span)


# ---------------------------------------------------------------------------
# Sheet 6 — Kontrol
# ---------------------------------------------------------------------------

def _sheet_control(sheet: Worksheet, period: str, stats: RunStats,
                   summaries: list[MonthSummary], anomalies: Collector,
                   settings: Settings, generated_at: datetime) -> None:
    span = 4
    styles.write_title(sheet, 1, "KONTROL — mutabakat ve varsayımlar", span)
    for index, width in enumerate([46, 22, 22, 60], start=1):
        from openpyxl.utils import get_column_letter
        sheet.column_dimensions[get_column_letter(index)].width = width

    row = 3

    def section(title: str) -> None:
        nonlocal row
        cell = sheet.cell(row=row, column=1, value=title)
        cell.font = styles.HEADER_FONT
        cell.fill = styles.HEADER_FILL
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        row += 1

    def line(label: str, value: object = "", note: str = "",
             fill=None) -> None:
        nonlocal row
        sheet.cell(row=row, column=1, value=label)
        cell = sheet.cell(row=row, column=2, value=value)
        cell.alignment = styles.RIGHT
        sheet.cell(row=row, column=4, value=note)
        if fill is not None:
            styles.style_row(sheet, row, span, fill)
        row += 1

    section("1. Okunan dosyalar")
    for name, filename in stats.files.items():
        line(_FILE_LABEL.get(name, name), filename)
    row += 1

    section("2. Satır mutabakatı")
    for name in ("roster", "macunkoy", "teknopark", "izin"):
        if name not in stats.rows_read:
            continue
        built = stats.records_built.get(name, 0)
        line(f"{_FILE_LABEL.get(name, name)} — okunan satır",
             stats.rows_read[name], f"kayda dönüşen: {built}")
    line("  İzin dosyasında atlanan ara toplam satırı",
         stats.rows_read.get("izin_ara_toplam_atlanan", 0),
         "Kişi başı ara toplam satırı — sayılsa izinler ikiye katlanırdı")
    line("  İzin dosyasından gelen uzaktan çalışma aralığı",
         stats.records_built.get("izin_uzaktan", 0),
         "Çalışma olarak sayıldı — uzaktan çalışma izin değil, çalışmadır")
    line("Ayıklanan ziyaretçi/geçici/stajyer kaydı", stats.excluded_badges,
         "Ziyaretçi / geçici / stajyer kartları — kişiye atfedilemediği için özetten düşer")
    outside = sum(stats.out_of_period.values())
    if outside or stats.out_of_period_leave:
        detail = ", ".join(f"{_FILE_LABEL.get(k, k)}: {v}"
                           for k, v in sorted(stats.out_of_period.items()))
        line(f"{period} dönemi dışında kalıp atılan kayıt", outside,
             detail or "—", styles.AMBER_FILL)
        line("  dönem dışı izin kaydı", stats.out_of_period_leave)
    else:
        line(f"{period} dönemi dışında kayıt", 0,
             "Tüm kayıtlar rapor dönemine ait", styles.GREEN_FILL)
    row += 1

    section("3. Dönem kapsamı")
    partial = [c for c in stats.coverage.values() if c.is_partial]
    for source, cov in sorted(stats.coverage.items()):
        note = "Dönemin tamamı kapsanıyor"
        fill = styles.GREEN_FILL
        if cov.is_partial:
            first = cov.trailing_missing[0].strftime("%d.%m.%Y")
            last = cov.trailing_missing[-1].strftime("%d.%m.%Y")
            note = (f"KISMİ DIŞA AKTARIM — {first} ve sonrası ({len(cov.trailing_missing)} "
                    f"iş günü) bu dosyada HİÇ YOK. Dosya ay bitmeden alınmış olabilir.")
            fill = styles.RED_FILL
        line(f"{_FILE_LABEL.get(source, source)} — kapsanan iş günü",
             f"{cov.present} / {cov.expected}", note, fill)
    # The per-source lines above only see a run of missing days at the END of the
    # period. A day on which NEITHER site recorded anybody is the one mid-period hole
    # that can be stated without a false alarm, and it has a second reading worth
    # printing: the day may simply be a holiday nobody marked. ADR-057.
    if stays := stats.blank_workdays:
        günler = ", ".join(d.strftime("%d.%m") for d in stays[:15])
        if len(stays) > 15:
            günler += f" ... (+{len(stays) - 15})"
        line("Hiçbir tesiste kaydı olmayan iş günü", len(stays),
             f"{günler}. Bu günler tatilse tatil listesine eklenmeli; değilse "
             f"kaynak dosyalarda o günler eksik.", styles.RED_FILL)
    else:
        line("Hiçbir tesiste kaydı olmayan iş günü", 0,
             "Beklenen her iş gününde en az bir tesiste kayıt var",
             styles.GREEN_FILL)
    if partial or stats.blank_workdays:
        line("SONUÇ", "RAPOR EKSİK",
             "Dönemin tamamı kapsanmıyor. Bu rapordaki saatler bordro için "
             "kullanılamaz — kaynak dosyalar ya da tatil listesi gözden geçirilmeli.",
             styles.RED_FILL)
    row += 1

    section("4. Hesaplama mutabakatı")
    computed = timedelta()
    for summary in summaries:
        computed += summary.gross
    gaps = stats.accepted_total - stats.union_total
    line("Kabul edilen aralık sayısı", stats.intervals_accepted)
    line("Kabul edilen aralıkların toplamı", hhmm(stats.union_total),
         "Sadece varlık süresi — gün içi boşluklar hariç")
    if settings.daily_hours == "union":
        line("Gün içi boşluklar", hhmm(gaps), "Sayılmadı — daily_hours: union")
    else:
        line("Gün içi boşluklar", hhmm(gaps),
             "Ödenen süreye dahil — gün içindeki boşluklar düşülmez", styles.GREY_FILL)
    line("Günlük ölçülen sürelerin toplamı", hhmm(stats.accepted_total),
         "Raporun ödediği süre")
    line("Kişi toplamlarının toplamı", hhmm(computed))
    matches = abs((computed - stats.accepted_total).total_seconds()) < 1
    line("Mutabakat", "TAMAM" if matches else "HATA",
         "Σ kişi = Σ günlük ölçülen süre" if matches
         else "Kayıt çift sayılmış veya kaybolmuş — inceleyin",
         styles.GREEN_FILL if matches else styles.RED_FILL)
    row += 1

    section("5. Kapsam")
    line("Raporda yer alan kişi", len(summaries))
    line("Mesai verisi olan", sum(1 for s in summaries if s.has_attendance))
    line("Mesai verisi olmayan", sum(1 for s in summaries if not s.has_attendance),
         "İzin kaydı var, kart kaydı yok. Bu kişilerin ayı eksik", styles.RED_FILL)
    line("Personel listesinde olmayan",
         sum(1 for s in summaries if not s.employee.in_roster),
         "Dönemde çalışıp personel listesi alınana kadar ayrılmış olabilir")
    # The other direction, and the one nothing used to say (ADR-071): people the roster
    # has and the period does not. They get no row, so every count above is blind to
    # them — which made them the one group a manual check could not reach.
    if stats.roster_only:
        line("Personel listesinde olup bu ayda hiç kaydı olmayan",
             len(stats.roster_only),
             f"Personel listesindeki {stats.roster_size} kişiden bu kadarının ne kart "
             "ne izin kaydı var, o yüzden rapora satır açılmıyor. Bir kısmı sonradan "
             "işe başlamış olabilir; listede işe giriş tarihi olmadığı için program "
             "ayırt edemiyor, tek tek bakılması gerekiyor",
             styles.RED_FILL)
        by_facility: dict[str, list[str]] = {}
        for name, facility in stats.roster_only:
            by_facility.setdefault(settings.facility(facility) or "—", []).append(name)
        for facility in sorted(by_facility, key=sort_key):
            names = by_facility[facility]
            line(f"  {facility}", len(names))
            for name in names:
                line(f"    {name}")
    line("Şüpheli kayıt (toplam)", len(anomalies))
    line("  toplama dahil edilmeyen",
         sum(1 for a in anomalies.items if a.severity == "excluded"))
    line("  bilgi amaçlı (sorun değil)",
         sum(1 for a in anomalies.items if a.severity == "info"),
         "Beklenen durum — kişilerin 'Şüpheli Kayıt' sayısına dahil edilmez, "
         "olağan durumlar", styles.GREY_FILL)
    row += 1

    section("6. Personel listesinde tekrarlanan kayıtlar")
    if stats.roster_duplicates:
        line("Aynı kişi için birden fazla hesap bulundu ve tekilleştirildi", "",
             "Aynı kontak no / e-posta, farklı kullanıcı adı. Eski hesap "
             "kapatılmamış olabilir.", styles.GREY_FILL)
        for note in stats.roster_duplicates:
            line(f"  {note}")
    else:
        line("Yok")
    row += 1

    section("7. Doğrulanmamış isim eşleştirmeleri")
    if settings.personnel.alias_pairs:
        line("Aşağıdaki eşleştirmeler UYGULANDI ama doğrulanmadı", "",
             "İki yazımın aynı kişi olduğu varsayıldı; yanlışsa iki kişinin "
             "saatleri birleşmiş olur", styles.AMBER_FILL)
        for variant, canonical in settings.personnel.alias_pairs:
            line(f"  {variant}", "->", canonical, styles.AMBER_FILL)
    else:
        line("Yok")
    row += 1

    # Same shape as the alias table above, and for the same reason: a mapping that
    # silently stopped applying looks exactly like a mapping that had nothing to do.
    # Listing what was seen and what it was shown as makes the difference visible.
    section("8. Tesis adları")
    seen = sorted({s.employee.facility for s in summaries if s.employee.facility},
                  key=sort_key)
    if seen:
        for raw in seen:
            shown = settings.facility(raw)
            if shown == raw:
                line(f"  {raw}", "->", "personel listesindeki hâliyle yazıldı",
                     styles.AMBER_FILL)
            else:
                line(f"  {raw}", "->", shown)
    else:
        line("Personel listesinde tesis bilgisi yok")
    row += 1

    section("9. Doğrulanmamış varsayımlar")
    # Only this month's. The calendar file holds every year the program has been run
    # for, so July's report listed May's seven holidays — dates outside the period it
    # reports on, presented as assumptions behind its figures.
    year, month = (int(part) for part in period.split("-"))
    in_period = sorted(day for day in settings.calendar.holidays
                       if (day.year, day.month) == (year, month))
    if in_period:
        line("Tatil günleri", f"{len(in_period)} gün",
             "Takvimde bu ay için işaretli günler")
        for day in in_period:
            line(f"  {day.strftime('%d.%m.%Y')}", _DAY_NAMES[day.weekday()])
    else:
        line("Tatil günleri", "bu ay için işaretli gün yok",
             "Varsa pencerenin Takvim ekranından işaretlenir", styles.AMBER_FILL)
    if settings.brk.deduct:
        line("Öğle arası", f"{settings.brk.minutes} dk",
             f"kalan mola kuralı, pencere {settings.brk.window_from:%H:%M}-"
             f"{settings.brk.window_to:%H:%M} arası")
    else:
        line("Öğle arası kesintisi", "YOK",
             "Mola için süre düşülmedi; kart süresi olduğu gibi sayıldı")
    line("Günlük süre ölçümü",
         "ilk giriş → son çıkış" if settings.daily_hours == "envelope"
         else "aralıkların toplamı",
         "Gün, ilk giriş ile son çıkış arasındaki süre olarak ölçülür"
         if settings.daily_hours == "envelope"
         else "Gün, giriş-çıkış aralıklarının toplamı olarak ölçülür")
    _roster_age_line(line, stats, period)
    row += 1

    section("10. Bu raporun kapsamadıkları")
    for text in (
        "Fazla mesai ve eksik çalışma hesabı",
        "Otomatik vardiya tespiti",
        "Multinet hakedişi",
        "Resmi tatil / hafta tatili çalışmasının ücret veya izin karşılığı",
        "Çalışanlara otomatik e-posta",
    ):
        line("  " + text)

    styles.write_footer(sheet, row + 1,
                        _footer_lines(period, stats, generated_at), span)


# ---------------------------------------------------------------------------

def _roster_age_line(line, stats: RunStats, period: str) -> None:
    """Report how stale the roster is relative to the reporting period.

    A roster newer than the period is why leavers and new hires appear (ADR-011).
    The date comes from the file's own timestamp — never hard-coded, or it silently
    lies the first time HR re-exports.
    """
    if stats.roster_date is None:
        line("Personel listesi tarihi", "bilinmiyor", "", styles.AMBER_FILL)
        return

    roster_month = stats.roster_date.year * 12 + stats.roster_date.month
    year, month = (int(part) for part in period.split("-"))
    gap = roster_month - (year * 12 + month)
    stamp = stats.roster_date.strftime("%d.%m.%Y")

    if gap == 0:
        line("Personel listesi tarihi", stamp,
             "Rapor dönemiyle aynı ay — ideal", styles.GREEN_FILL)
    elif gap > 0:
        line("Personel listesi tarihi", stamp,
             f"Rapor döneminden {gap} ay SONRA alınmış. Ayrılanlar listede "
             f"görünmez, sonradan girenler mesai verisinde yoktur",
             styles.AMBER_FILL)
    else:
        line("Personel listesi tarihi", stamp,
             f"Rapor döneminden {-gap} ay ÖNCE alınmış. O tarihten sonra işe "
             f"girenler listede yok", styles.AMBER_FILL)


def _tags_label(tags: frozenset[str]) -> str:
    """A day's tags in words, in a stable order.

    Sorted by the display text rather than by the internal name, so the column reads
    alphabetically to somebody who cannot see the internal names. An unknown tag is
    printed as it stands: tags come from `merge.py`, and a silently dropped one would
    be a fact removed from the audit trail.
    """
    return ", ".join(sorted(TAG_TEXT.get(tag, tag) for tag in tags))


def _day_sources_label(workday: WorkDay) -> str:
    """Where the day began and where it ended, when those are different places.

    `Macunköy → Teknopark` reads as one sentence and needs no second column. The plain
    site name is kept for the ordinary case, and `+` for a day whose first or last
    interval was itself built from both sites' records — there the question has no
    single answer, and inventing one would be worse than saying so.

    Measured on July 2026: 2 437 person-days at one site, 286 with a merged interval at
    one end, and 8 that genuinely start at one site and end at the other.
    """
    if not workday.intervals:
        return ""
    first = _sources_label(workday.intervals[0].sources)
    last = _sources_label(workday.intervals[-1].sources)
    if first and last and first != last and "+" not in first and "+" not in last:
        return f"{first} → {last}"
    return _sources_label(
        frozenset().union(*(iv.sources for iv in workday.intervals)))


def _sources_label(sources: frozenset[str]) -> str:
    real = sorted(_SOURCE_LABEL.get(s, s) for s in sources if s != "uzaktan")
    if "uzaktan" in sources or "izin" in sources:
        if "Uzaktan" not in real:
            real.append("Uzaktan")
    if len(real) > 1:
        return " + ".join(real)
    return real[0] if real else ""


def _period_label(period: str) -> str:
    months = ("OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN", "TEMMUZ",
              "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK")
    year, month = period.split("-")
    return f"{months[int(month) - 1]} {year}"


def _footer_lines(period: str, stats: RunStats, generated_at: datetime) -> list[str]:
    files = ", ".join(stats.files.values())
    return [
        f"Oluşturulma: {generated_at:%d.%m.%Y %H:%M} · Dönem: {period} · "
        f"mesai-takip v0.1.0",
        f"Kaynak dosyalar: {files}",
        "Bu rapor otomatik üretilmiştir; aynı dosyalarla her zaman aynı sonucu verir.",
    ]
