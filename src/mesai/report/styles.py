"""All report styling in one place.

Colour is never the only signal — every shaded row also carries text, so a
colour-blind reader loses nothing.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

TITLE_FONT = Font(bold=True, size=14, color="1F3864")
BANNER_FONT = Font(italic=True, size=10, color="7F4F00")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
FOOTER_FONT = Font(italic=True, size=9, color="808080")
TOTAL_FONT = Font(bold=True, size=10)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
AMBER_FILL = PatternFill("solid", fgColor="FFF2CC")     # attention
RED_FILL = PatternFill("solid", fgColor="FCE4E4")       # excluded from totals
GREY_FILL = PatternFill("solid", fgColor="F2F2F2")      # informational
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")     # reconciled
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")

_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_header(sheet: Worksheet, row: int, headers: list[str],
                 widths: list[int]) -> None:
    for index, (label, width) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(row=row, column=index, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[row].height = 30
    sheet.freeze_panes = sheet.cell(row=row + 1, column=1)
    sheet.auto_filter.ref = (
        f"A{row}:{get_column_letter(len(headers))}{row}"
    )


def write_title(sheet: Worksheet, row: int, text: str, span: int) -> None:
    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def write_banner(sheet: Worksheet, row: int, text: str, span: int) -> None:
    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = BANNER_FONT
    cell.fill = AMBER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    sheet.row_dimensions[row].height = 28


def write_footer(sheet: Worksheet, row: int, lines: list[str], span: int) -> None:
    for offset, line in enumerate(lines):
        cell = sheet.cell(row=row + offset, column=1, value=line)
        cell.font = FOOTER_FONT
        sheet.merge_cells(start_row=row + offset, start_column=1,
                          end_row=row + offset, end_column=span)


def style_row(sheet: Worksheet, row: int, columns: int,
              fill: PatternFill | None = None) -> None:
    for index in range(1, columns + 1):
        cell = sheet.cell(row=row, column=index)
        cell.border = BORDER
        if fill is not None:
            cell.fill = fill
