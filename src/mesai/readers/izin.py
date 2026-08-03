"""HCM leave export — `HCMT34_*_IZIN.xlsx`.

Two things this reader must get right:

1. **Skip the per-person subtotal row.** The first row for each person has the leave
   type empty and holds the sum of their later rows in `Kullanılan Gün`. Counting it
   doubles every person's leave. (docs/DATA-SOURCES.md D6)

2. **`Uzaktan Çalışma` is worked time, not leave.** Those rows carry real start and
   end times and become PunchRecords that enter the same interval union as badge
   records. Every other type is absence. (ADR-007)
"""

from __future__ import annotations

import warnings
from datetime import datetime, time, timedelta
from pathlib import Path

import openpyxl

from ..anomalies import Anomaly, AnomalyKind
from ..config import Settings
from ..models import LeaveRecord, PunchRecord
from ..normalize import name_key
from .base import LayoutError, as_date, as_text, as_time, clean_name

SOURCE = "izin"
SHEET_CANDIDATES = ("HCMPERS",)
EXPECTED_HEADERS = ("Sicil No", "Görünen Ad", "İzin Tipi", "Kullanılan Gün")

_COL_BADGE, _COL_NAME, _COL_DEPARTMENT = 0, 1, 2
_COL_TYPE, _COL_STATUS_APPROVAL, _COL_STATUS = 4, 5, 6
_COL_START_DATE, _COL_START_TIME, _COL_END_DATE, _COL_END_TIME = 7, 8, 9, 10
_COL_DAYS = 15


def read(
    path: Path, settings: Settings
) -> tuple[list[LeaveRecord], list[PunchRecord], list[Anomaly], int, int]:
    """Returns (leave_records, remote_work_punches, anomalies, rows_read, subtotals_skipped)."""
    # The HCM export carries no default cell style, so openpyxl warns on every load.
    # Expected and harmless — openpyxl substitutes its own default and we read values
    # only. Suppressed narrowly so real warnings still surface.
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore", message="Workbook contains no default style", category=UserWarning)
        workbook = openpyxl.load_workbook(path, data_only=True)

    sheet_name = next((s for s in SHEET_CANDIDATES if s in workbook.sheetnames),
                      workbook.sheetnames[0])
    sheet = workbook[sheet_name]

    rows = sheet.iter_rows(values_only=True)
    header = [as_text(c) for c in next(rows)]
    missing = [h for h in EXPECTED_HEADERS if h not in header]
    if missing:
        workbook.close()
        raise LayoutError(f"{path.name}: beklenen kolonlar yok: {missing}")

    leave: list[LeaveRecord] = []
    remote: list[PunchRecord] = []
    anomalies: list[Anomaly] = []
    rows_read = 0
    subtotals = 0

    for row_no, row in enumerate(rows, start=2):
        if row is None or all(c is None for c in row):
            continue
        name = as_text(row[_COL_NAME])
        if not name:
            continue
        rows_read += 1

        leave_type = as_text(row[_COL_TYPE])
        if not leave_type:
            subtotals += 1          # D6 — per-person subtotal row
            continue

        display = clean_name(name)
        key = settings.personnel.resolve(name_key(display))
        start = _combine(row[_COL_START_DATE], row[_COL_START_TIME])
        end = _combine(row[_COL_END_DATE], row[_COL_END_TIME])

        leave.append(LeaveRecord(
            key=key,
            raw_name=display,
            personnel_no=as_text(row[_COL_BADGE]),
            leave_type=leave_type,
            status=as_text(row[_COL_STATUS]),
            start=start,
            end=end,
            days=_as_float(row[_COL_DAYS]),
            department=as_text(row[_COL_DEPARTMENT]),
            source_row=row_no,
        ))

        if leave_type in settings.worked_leave_types:
            punches, notes = _remote_punches(
                key, display, start, end, row_no, settings)
            remote.extend(punches)
            anomalies.extend(notes)

    workbook.close()
    return leave, remote, anomalies, rows_read, subtotals


def _combine(date_value: object, time_value: object) -> datetime | None:
    day = as_date(date_value)
    if day is None:
        return None
    clock = as_time(time_value) or time()
    return datetime.combine(day, clock)


def _as_float(value: object) -> float:
    try:
        return float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return 0.0


def _remote_punches(
    key: tuple[str, str], display: str, start: datetime | None, end: datetime | None,
    row_no: int, settings: Settings,
) -> tuple[list[PunchRecord], list[Anomaly]]:
    """Turn one `Uzaktan Çalışma` row into per-day worked intervals."""
    if start is None or end is None or end <= start:
        return [], [Anomaly(
            kind=AnomalyKind.UNPARSEABLE_ROW, source=SOURCE, source_row=row_no,
            key=key, raw_name=display,
            detail="uzaktan çalışma kaydında geçerli başlangıç/bitiş yok",
        )]

    if start.date() == end.date():
        return [PunchRecord(
            source=SOURCE, source_row=row_no, raw_name=display, key=key,
            date=start.date(), entry=start, exit=end, tag="uzaktan",
        )], []

    # Multi-day row: split into one interval per expected working day using the
    # standard shift window. Flagged, because the source does not state daily hours.
    punches: list[PunchRecord] = []
    day = start.date()
    while day <= end.date():
        if not settings.calendar.is_rest_day(day) and not settings.calendar.is_holiday(day):
            day_start = start if day == start.date() else datetime.combine(
                day, settings.shift_start)
            day_end = end if day == end.date() else datetime.combine(
                day, settings.shift_end)
            if day_end > day_start:
                punches.append(PunchRecord(
                    source=SOURCE, source_row=row_no, raw_name=display, key=key,
                    date=day, entry=day_start, exit=day_end, tag="uzaktan",
                ))
        day += timedelta(days=1)

    note = Anomaly(
        kind=AnomalyKind.MULTI_DAY_REMOTE, source=SOURCE, source_row=row_no,
        key=key, raw_name=display, date=start.date(),
        raw_entry=start.strftime("%d.%m.%Y %H:%M"),
        raw_exit=end.strftime("%d.%m.%Y %H:%M"),
        detail=f"{len(punches)} iş gününe bölündü, günlük saatler vardiyadan alındı",
    )
    return punches, [note]
