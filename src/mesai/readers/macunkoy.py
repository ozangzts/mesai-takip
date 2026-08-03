"""Macunköy badge-terminal export — flat log, one row per person-day.

Layout (docs/DATA-SOURCES.md §1):
    A Ad | B Soyad | C Personel | D SicilNo | E Birim | F (empty) | G Bolum
    H MesaiTarih | I Giris | J Cikis | K SureSaat

Health: poor. 388 of 1 209 rows are missing a punch, 29 have a negative duration,
23 identities are visitor/temporary badges. None of that is fixed here — the reader
passes it through and the rules layer decides.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from ..anomalies import Anomaly, AnomalyKind
from ..config import Settings
from ..models import PunchRecord
from ..normalize import fold, is_excluded, name_key
from .base import LayoutError, as_date, as_datetime, as_text, clean_name

SOURCE = "macunkoy"
EXPECTED_HEADERS = ("Ad", "Soyad", "Personel", "MesaiTarih", "Giris", "Cikis")

_COL_GIVEN, _COL_SURNAME, _COL_FULL, _COL_BADGE = 0, 1, 2, 3
_COL_DEPARTMENT, _COL_DATE, _COL_ENTRY, _COL_EXIT, _COL_DURATION = 6, 7, 8, 9, 10


def read(path: Path, settings: Settings) -> tuple[list[PunchRecord], list[Anomaly], int, int]:
    """Returns (records, anomalies, rows_read, excluded_badge_rows)."""
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows = sheet.iter_rows(values_only=True)
    header = [as_text(c) for c in next(rows)]
    missing = [h for h in EXPECTED_HEADERS if h not in header]
    if missing:
        workbook.close()
        raise LayoutError(
            f"{path.name}: beklenen kolonlar yok: {missing} — başlık: {header}"
        )

    records: list[PunchRecord] = []
    anomalies: list[Anomaly] = []
    rows_read = 0
    excluded = 0

    for row_no, row in enumerate(rows, start=2):
        if row is None or all(c is None for c in row):
            continue
        rows_read += 1

        full = as_text(row[_COL_FULL])
        given = as_text(row[_COL_GIVEN])
        surname = as_text(row[_COL_SURNAME])
        if not full and not given:
            anomalies.append(Anomaly(
                kind=AnomalyKind.UNPARSEABLE_ROW, source=SOURCE, source_row=row_no,
                detail="isim kolonu boş",
            ))
            continue
        display = clean_name(full or f"{given} {surname or ''}")

        if is_excluded(display, given, surname, settings.personnel.exclude_prefixes):
            excluded += 1
            continue

        day = as_date(row[_COL_DATE])
        if day is None:
            anomalies.append(Anomaly(
                kind=AnomalyKind.UNPARSEABLE_ROW, source=SOURCE, source_row=row_no,
                raw_name=display, detail=f"tarih okunamadı: {row[_COL_DATE]!r}",
            ))
            continue

        records.append(PunchRecord(
            source=SOURCE,
            source_row=row_no,
            raw_name=display,
            key=settings.personnel.resolve(name_key(display)),
            date=day,
            entry=as_datetime(row[_COL_ENTRY]),
            exit=as_datetime(row[_COL_EXIT]),
            badge_id=_badge(row[_COL_BADGE]),
            department=as_text(row[_COL_DEPARTMENT]),
            reported_duration=as_text(row[_COL_DURATION]),
        ))

    workbook.close()
    return records, anomalies, rows_read, excluded


def _badge(value: object) -> str | None:
    """Keep only numeric personnel numbers. `SN`-prefixed values are card numbers
    that do not correspond to anything else (docs/DATA-SOURCES.md D5) and must
    never reach the report."""
    text = as_text(value)
    if not text:
        return None
    if fold(text).startswith("SN"):
        return None
    return text
